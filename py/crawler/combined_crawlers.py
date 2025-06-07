import datetime
import json
import logging
import re
from io import BytesIO
from tempfile import TemporaryFile
from typing import Any, List
from typing import Optional

from csv import DictWriter
from decimal import Decimal
from logging import getLogger
from os import makedirs
from pathlib import Path
from zipfile import ZipFile, ZIP_DEFLATED

from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from typing import Optional, overload

import os
from typing import List

import urllib.parse

from bs4 import BeautifulSoup
from crawler.store.models import Product, Store

import openpyxl
from crawler.store.models import Product, Store


from pydantic import BaseModel, Field


# ============================================================================
# BASE MODELS CLASS
# ============================================================================

class Product(BaseModel):
    """
    Unified product model for all stores.
    """

    product: str  # Product name
    product_id: str  # Store specific product identifier
    brand: str  # Brand name
    quantity: str  # Amount (e.g., "500g", "1L")
    unit: str  # Unit of measure (e.g., "kg", "kom")
    price: Decimal  # Current retail price
    unit_price: Decimal  # Price per unit of measure
    barcode: str  # EAN/barcode
    category: str  # Product category

    # Optional fields that appear in some stores
    best_price_30: Optional[Decimal] = None  # Lowest price in last 30 days
    special_price: Optional[Decimal] = None  # Promotional/discounted price
    anchor_price: Optional[Decimal] = None  # Reference price (often May 2, 2025)
    anchor_price_date: Optional[str] = None  # Date of reference price
    packaging: Optional[str] = None  # Packaging information
    initial_price: Optional[Decimal] = (
        None  # Initial price for newly added products (if available)
    )
    date_added: Optional[date] = None  # When the product was added (if available)

    def __str__(self):
        return f"{self.brand.title()} {self.product.title()} (EAN: {self.barcode})"


class Store(BaseModel):
    """
    Unified store model for all retailers.
    """

    chain: str  # Store chain name, lowercase ("konzum", "lidl", "spar", etc.)
    store_id: str  # Chain-specific store (location) identifier
    name: str  # Store name (e.g., "Lidl Zagreb")
    store_type: str  # Type (e.g., "supermarket", "hipermarket")
    city: str  # City location
    street_address: str  # Street address
    zipcode: str = ""  # Postal code (empty default if not available)
    items: List[Product] = Field(default_factory=list)  # Products in this store

    def __str__(self):
        return f"{self.name} ({self.street_address})"


# ============================================================================
# OUTPUT CLASS
# ============================================================================

logger = getLogger(__name__)

STORE_COLUMNS = [
    "store_id",
    "type",
    "address",
    "city",
    "zipcode",
]

PRODUCT_COLUMNS = [
    "product_id",
    "barcode",
    "name",
    "brand",
    "category",
    "unit",
    "quantity",
]

PRICE_COLUMNS = [
    "store_id",
    "product_id",
    "price",
    "unit_price",
    "best_price_30",
    "anchor_price",
    "special_price",
]


def transform_products(
    stores: list[Store],
) -> tuple[list[dict], list[dict], list[dict]]:
    """
    Transform store data into a structured format for CSV export.

    Args:
        stores: List of Store objects containing product data.

    Returns:
        Tuple containing:
            - List of store dictionaries with STORE_COLUMNS
            - List of product dictionaries with PRODUCT_COLUMNS
            - List of price dictionaries with PRICE_COLUMNS
    """
    store_list = []
    product_map = {}
    price_list = []

    def maybe(val: Decimal | None) -> Decimal | str:
        return val if val is not None else ""

    for store in stores:
        store_data = {
            "store_id": store.store_id,
            "type": store.store_type,
            "address": store.street_address,
            "city": store.city,
            "zipcode": store.zipcode or "",
        }
        store_list.append(store_data)

        for product in store.items:
            key = f"{store.chain}:{product.product_id}"
            if key not in product_map:
                product_map[key] = {
                    "barcode": product.barcode or key,
                    "product_id": product.product_id,
                    "name": product.product,
                    "brand": product.brand,
                    "category": product.category,
                    "unit": product.unit,
                    "quantity": product.quantity,
                }
            price_list.append(
                {
                    "store_id": store.store_id,
                    "product_id": product.product_id,
                    "price": product.price,
                    "unit_price": maybe(product.unit_price),
                    "best_price_30": maybe(product.best_price_30),
                    "anchor_price": maybe(product.anchor_price),
                    "special_price": maybe(product.special_price),
                }
            )

    return store_list, list(product_map.values()), price_list


def save_csv(path: Path, data: list[dict], columns: list[str]):
    """
    Save data to a CSV file.

    Args:
        path: Path to the CSV file.
        data: List of dictionaries containing the data to save.
        columns: List of column names for the CSV file.
    """
    if not data:
        logger.warning(f"No data to save at {path}, skipping")
        return

    if set(columns) != set(data[0].keys()):
        raise ValueError(
            f"Column mismatch: expected {columns}, got {list(data[0].keys())}"
        )
        return

    with open(path, "w", newline="") as f:
        writer = DictWriter(f, fieldnames=columns)
        writer.writeheader()
        for row in data:
            writer.writerow({k: str(v) for k, v in row.items()})


def save_chain(chain_path: Path, stores: list[Store]):
    """
    Save retail chain data to CSV files.

    This function creates a directory for the retail chain and saves:

    * stores.csv - containing store information with STORE_COLUMNS
    * products.csv - containing product information with PRODUCT_COLUMNS
    * prices.csv - containing price information with PRICE_COLUMNS

    Args:
        chain_path: Path to the directory where CSV files will be saved
            (will be created if it doesn't exist).
        stores: List of Store objects containing product data.
    """

    makedirs(chain_path, exist_ok=True)
    store_list, product_list, price_list = transform_products(stores)
    save_csv(chain_path / "stores.csv", store_list, STORE_COLUMNS)
    save_csv(chain_path / "products.csv", product_list, PRODUCT_COLUMNS)
    save_csv(chain_path / "prices.csv", price_list, PRICE_COLUMNS)


def copy_archive_info(path: Path):
    archive_info = open(Path(__file__).parent / "archive-info.txt", "r").read()
    with open(path / "archive-info.txt", "w") as f:
        f.write(archive_info)


def create_archive(path: Path, output: Path):
    """
    Create a ZIP archive of price files for a given date.

    Args:
        path: Path to the directory to archive.
        output: Path to the output ZIP file.
    """
    with ZipFile(output, "w", compression=ZIP_DEFLATED, compresslevel=9) as zf:
        for file in path.rglob("*"):
            zf.write(file, arcname=file.relative_to(path))


# ============================================================================
# BASE CRAWLER CLASS
# ============================================================================

class BaseCrawler:
    """
    Base crawler class with common functionality and interface for all crawlers.
    """

    CHAIN: str
    BASE_URL: str

    TIMEOUT = 30.0
    USER_AGENT = None

    ZIP_DATE_PATTERN: re.Pattern | None = None

    PRICE_MAP: dict[str, tuple[str, bool]]
    """Mapping from CSV column names to price fields and whether they are required."""

    FIELD_MAP: dict[str, tuple[str, bool]]
    """Mapping from CSV column names to non-price fields and whether they are required."""

    def __init__(self):
        self.client = httpx.Client(timeout=30.0, follow_redirects=True)

    def fetch_text(
        self,
        url: str,
        encodings: list[str] | None = None,
        prefix: str | None = None,
    ) -> str:
        """
        Download a text file (web page or CSV) from the given URL.

        Args:
            url: URL to download from
            encoding: Optional encoding to decode the content. If None, uses default.

        Returns:
            The content of the file as a string, or an empty string if the download fails.
        """

        def try_decode(content: bytes) -> str:
            for encoding in encodings:  # type: ignore
                try:
                    text = content.decode(encoding)
                    if not prefix or text.startswith(prefix):
                        return text
                except UnicodeDecodeError:
                    continue
            raise ValueError(f"Error decoding {url} - tried: {encodings}")

        logger.debug(f"Fetching {url}")
        try:
            response = self.client.get(url)
            response.raise_for_status()
            if encodings:
                return try_decode(response.content)
            else:
                return response.text
        except httpx.RequestError as e:
            logger.error(f"Download from {url} failed: {e}", exc_info=True)
            raise

    def fetch_binary(self, url: str, fp: BinaryIO):
        """
        Download a binary file to a provided location.

        The location should be created using tempfile.NamedTemporaryFile

        Args:
            url: URL of the ZIP file to download

        Returns:
            Path to the downloaded ZIP file
        """

        logger.info(f"Downloading binary file from {url}")

        MB = 1024 * 1024

        t0 = time()
        with self.client.stream("GET", url) as response:
            response.raise_for_status()
            total_mb = int(response.headers.get("content-length", 0)) // MB
            logger.debug(f"File size: {total_mb} MB")

            for chunk in response.iter_bytes(chunk_size=1 * MB):
                fp.write(chunk)

        t1 = time()
        dt = int(t1 - t0)
        logger.debug(f"Downloaded {total_mb} MB in {dt}s")

    def read_csv(self, text: str, delimiter: str = ",") -> DictReader:
        return DictReader(text.splitlines(), delimiter=delimiter)  # type: ignore

    def get_zip_contents(
        self, url: str, suffix: str
    ) -> Generator[tuple[str, bytes], None, None]:
        with NamedTemporaryFile(mode="w+b") as temp_zip:
            self.fetch_binary(url, temp_zip)  # type: ignore
            temp_zip.seek(0)

            with ZipFile(temp_zip, "r") as zip_fp:
                for file_info in zip_fp.infolist():
                    if not file_info.filename.endswith(suffix):
                        continue

                    logger.debug(f"Processing file: {file_info.filename}")

                    try:
                        with zip_fp.open(file_info) as file:
                            xml_content = file.read()
                            yield (file_info.filename, xml_content)
                    except Exception as e:
                        logger.error(
                            f"Error processing file {file_info.filename}: {e}",
                            exc_info=True,
                        )

    @staticmethod
    def parse_price(
        price_str: str | None,
        required: bool = True,
    ) -> Decimal | None:
        """
        Parse a price string using the global parse_price function.
        """
        return parse_price(price_str, required)

    @staticmethod
    def strip_diacritics(text: str) -> str:
        """
        Remove diacritics from a string.

        Args:
            text: The input string

        Returns:
            The string with diacritics removed
        """
        return "".join(
            c
            for c in unicodedata.normalize("NFD", text)
            if unicodedata.category(c) != "Mn"
        )

    def fix_product_data(self, data: dict[str, Any]) -> dict[str, Any]:
        """
        Do any cleaning or transformation of the Product data here.

        Args:
            data: Dictionary containing the row data

        Returns:
            The cleaned or transformed data
        """
        # Common fixups for all crawlers
        if data["barcode"] == "":
            data["barcode"] = f"{self.CHAIN}:{data['product_id']}"
        data["barcode"] = data["barcode"].replace('"', "").replace("'", "").strip()

        if "special_price" not in data:
            data["special_price"] = None

        if data["price"] is None:
            if data.get("special_price") is None:
                if data.get("unit_price") is not None:
                    data["price"] = data["unit_price"]
                else:
                    raise ValueError(
                        "Price, special price, and unit price are all missing"
                    )
            else:
                data["price"] = data["special_price"]

        if data["anchor_price"] is not None and not data.get("anchor_price_date"):
            data["anchor_price_date"] = datetime.date(2025, 5, 2).isoformat()

        if data["unit_price"] is None:
            data["unit_price"] = data["price"]

        return data

    def parse_csv_row(self, row: dict) -> Product:
        """
        Parse a single row of CSV data into a Product object.
        """
        data = {}

        for field, (column, is_required) in self.PRICE_MAP.items():
            value = row.get(column)
            try:
                data[field] = self.parse_price(value, is_required)
            except ValueError as err:
                logger.warning(
                    f"Failed to parse {field} from {column}: {err}",
                    exc_info=True,
                )
                raise

        for field, (column, is_required) in self.FIELD_MAP.items():
            value = row.get(column, "").strip()
            if not value and is_required:
                raise ValueError(f"Missing required field: {field}")
            data[field] = value

        data = self.fix_product_data(data)
        return Product(**data)  # type: ignore

    def parse_xml_product(self, elem: Any) -> Product:
        def get_text(xpath: Any, default=""):
            elements = elem.xpath(xpath)
            return elements[0] if elements and elements[0] else default

        data = {}
        for field, (tagname, is_required) in self.PRICE_MAP.items():
            value = get_text(f"{tagname}/text()")
            try:
                data[field] = self.parse_price(value, is_required)
            except ValueError as err:
                logger.warning(
                    f"Failed to parse {field} from {tagname}: {err}",
                    exc_info=True,
                )
                raise

        for field, (tagname, is_required) in self.FIELD_MAP.items():
            value = get_text(f"{tagname}/text()")
            if not value and is_required:
                raise ValueError(
                    f"Missing required field: {field} (expected <{tagname}>)"
                )
            data[field] = value

        data = self.fix_product_data(data)
        return Product(**data)  # type: ignore

    def parse_csv(self, content: str, delimiter: str = ",") -> list[Product]:
        """
        Parses CSV content into Product objects.

        Args:
            content: CSV content as a string
            delimiter: Delimiter used in the CSV file (default: ",")

        Returns:
            List of Product objects
        """
        logger.debug("Parsing CSV content")

        products = []
        for row in self.read_csv(content, delimiter=delimiter):
            try:
                product = self.parse_csv_row(row)
            except Exception as e:
                logger.warning(f"Failed to parse row: {row}: {e}")
                continue
            products.append(product)

        logger.debug(f"Parsed {len(products)} products from CSV")
        return products

    def parse_index_for_zip(self, html_content: str) -> dict[datetime.date, str]:
        """
        Parse HTML and return ZIP links.

        Args:
            html_content: HTML content of the price list index page

        Returns:
            Dictionary mapping dates to ZIP file URLs
        """

        if not self.ZIP_DATE_PATTERN:
            raise NotImplementedError(
                f"{self.__class__.__name__}.ZIP_DATE_PATTERN is not defined"
            )

        soup = BeautifulSoup(html_content, "html.parser")
        zip_urls_by_date = {}

        links = soup.select('a[href$=".zip"]')
        for link in links:
            url = str(link["href"])

            m = self.ZIP_DATE_PATTERN.match(url)
            if not m:
                continue

            # Extract date from the URL
            day, month, year = m.groups()
            url_date = datetime.date(int(year), int(month), int(day))
            zip_urls_by_date[url_date] = url

        return zip_urls_by_date

    def get_all_products(self, date: datetime.date) -> list[Store]:
        raise NotImplementedError()

    def crawl(self, date: datetime.date) -> list[Store]:
        name = self.CHAIN.capitalize()
        logger.info(f"Starting {name} crawl for date: {date}")
        t0 = time()

        try:
            stores = self.get_all_products(date)
            n_prices = sum(len(store.items) for store in stores)

            t1 = time()
            dt = int(t1 - t0)

            logger.info(
                f"Completed {name} crawl for {date} in {dt}s, "
                f"found {len(stores)} stores with {n_prices} total prices"
            )
            return stores

        except Exception as e:
            logger.error(f"Error crawling {name} price list: {e}", exc_info=True)
            raise
            

# ============================================================================
# UTILS
# ============================================================================

logger = logging.getLogger(__name__)


def to_camel_case(text: str) -> str:
    """
    Converts text to camel case and replace any '_' with ' '.

    Args:
        text: Input text

    Returns:
        Text converted to camel case
    """
    if text:
        return text.replace("_", " ").title()
    else:
        return ""


@overload
def parse_price(price_str: str, required: bool = True) -> Decimal: ...


@overload
def parse_price(price_str: str, required: bool = False) -> Decimal | None: ...


def parse_price(price_str: str | None, required: bool = False) -> Decimal | None:
    """
    Parse a price string.

    The string may use either , or . as decimal separator, may omit leading
    zero, and may contain currency symbols "€" or "EUR".

    None is handled the same as empty string - no price information available.

    Args:
        price_str: String representing the price, or None (no price)
        required: If True, raises ValueError if the price is not valid
                  If False, returns None for invalid prices

    Returns:
        Parsed price as a Decimal with 2 decimal places

    Raises:
        ValueError: If required is True and the price is not valid
    """
    if price_str is None:
        price_str = ""

    price_str = price_str.replace("€", "").replace("EUR", "").replace(",", ".").strip()

    if not price_str:
        if required:
            raise ValueError("Price is required")
        else:
            return None

    # Handle missing leading zero
    if price_str.startswith("."):
        price_str = "0" + price_str

    try:
        # Convert to Decimal and round to 2 decimal places
        return Decimal(price_str).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except (ValueError, TypeError, InvalidOperation):
        logger.warning(f"Failed to parse price: {price_str}")
        if required:
            raise ValueError(f"Invalid price format: {price_str}")
        else:
            return None


def log_operation_timing(
    operation_name: str,
    store_name: str,
    date: datetime.date,
    start_time: float,
    end_time: float,
    store_count: int,
    total_products: int,
) -> None:
    """
    Log the timing information for a crawler operation.

    Args:
        operation_name: Name of the operation being timed
        store_name: Name of the store being crawled
        date: The date for which the crawl was performed
        start_time: The start time in seconds
        end_time: The end time in seconds
        store_count: The number of stores processed
        total_products: The total number of products found
    """
    dt = int(end_time - start_time)
    logger.info(
        f"Completed {store_name} {operation_name} for {date} in {dt}s, "
        f"found {store_count} stores with {total_products} total products"
    )


def extract_zipcode_from_text(text: str) -> Optional[str]:
    """
    Extracts a zipcode (postal code) from text using a regex pattern.

    Args:
        text: Text that might contain a zipcode

    Returns:
        The extracted zipcode or None if not found
    """
    # Common pattern for Croatian zipcodes (5 digits)
    zipcode_pattern = r"\b(\d{5})\b"
    match = re.search(zipcode_pattern, text)
    return match.group(1) if match else None


# ============================================================================
# DM CRAWLER CLASSES
# ============================================================================


logger = logging.getLogger(__name__)


class DmCrawler(BaseCrawler):
    """Crawler for DM (DrogerieMarkt) store prices."""

    CHAIN = "dm"
    BASE_URL = "https://www.dm.hr"
    CONTENT_BASE_URL = "https://content.services.dmtech.com/rootpage-dm-shop-hr-hr"
    INDEX_URL = f"{CONTENT_BASE_URL}/novo/promocije/nove-oznake-cijena-i-vazeci-cjenik-u-dm-u-2906632?mrclx=false"

    # DM has global prices, not per-store prices
    STORE_ID = "all"
    STORE_NAME = "DM"

    def parse_date_from_title(self, title: str) -> datetime.date:
        """
        Extract date from the title the Excel link.

        Args:
            title: Title attribute of the link

        Returns:
            Extracted date object
        """
        # Match date in format DD.MM.YYYY where D or M can be single-digit
        date_match = re.search(r"(\d{1,2})\.(\d{1,2})\.(\d{4})", title)
        if not date_match:
            raise ValueError(f"Could not extract date from title: {title}")

        day, month, year = map(int, date_match.groups())
        return datetime.date(year, month, day)

    def find_excel_url(self, json_content: str, target_date: datetime.date) -> str:
        """
        Parse the JSON data to find the Excel file URL for the target date.

        Args:
            json_content: JSON content from the index page
            target_date: The date to search for

        Returns:
            URL of the Excel file

        Raises:
            ValueError: If no Excel file is found for the target date
        """
        try:
            # Parse JSON data
            data = json.loads(json_content)
        except json.JSONDecodeError as e:
            logger.error(f"Failed to parse JSON: {e}")
            raise ValueError("Failed to parse JSON data from the page")

        # Find all CMDownload entries in mainData
        excel_entries = []
        for item in data.get("mainData", []):
            if item.get("type") == "CMDownload":
                excel_entries.append(item.get("data", {}))

        if not excel_entries:
            logger.warning("No Excel links found in JSON data")
            raise ValueError("No Excel links found in JSON data")

        target_date_str = f"{target_date.day}.{target_date.month}.{target_date.year}"
        logger.info(f"Looking for Excel file with date {target_date_str}")

        for entry in excel_entries:
            headline = entry.get("headline", "")
            link_target = entry.get("linkTarget", "")

            if not headline or not link_target:
                continue

            try:
                link_date = self.parse_date_from_title(headline)
                if link_date == target_date:
                    # Ensure URL is absolute
                    if not link_target.startswith(("http://", "https://")):
                        url = f"{self.CONTENT_BASE_URL}{link_target}"
                    else:
                        url = link_target
                    logger.info(f"Found Excel file with date {link_date}: {url}")
                    return url
            except Exception as e:
                logger.warning(f"Error parsing date from headline '{headline}': {e}")
                continue

        raise ValueError(f"No Excel file found for date {target_date_str}")

    def detect_columns(self, worksheet: Any) -> list[str]:
        """
        Detect the column ordering in the DM Excel worksheet.

        This relies on the fact that one of the columns in the header will
        always be "naziv + šifra", which is a merged cell that actually
        has two cells in the data, naziv and product ID.

        Args:
            worksheet: The active worksheet object

        Returns:
            List of column headers
        """
        for row in worksheet.iter_rows():
            row_str = [
                self.strip_diacritics(str(cell.value or "").lower()) for cell in row
            ]
            if "naziv + sifra" in row_str:
                idx = row_str.index("naziv + sifra")
                if row_str[idx + 1] != "":
                    raise ValueError(
                        "Expected 'naziv + šifra' to be a merged cell with two parts"
                    )
                row_str[idx] = "naziv"
                row_str[idx + 1] = "sifra"
                return row_str

        raise ValueError(
            "Could not detect Excel columns, DM file format may have changed"
        )

    @staticmethod
    def map_columns(row: Any, columns: list) -> dict[str, Any]:
        """
        Map the row data to a dictionary using the detected columns.

        Args:
            row: The row object from the worksheet
            columns: List of column headers

        Returns:
            Dictionary mapping column names to cell values
        """
        return {col: str(row[i].value or "").strip() for i, col in enumerate(columns)}

    def parse_excel(self, excel_data: bytes) -> List[Product]:
        """
        Parse Excel file data into Product objects.

        Args:
            excel_data: Raw Excel file content

        Returns:
            List of Product objects
        """
        logger.debug("Parsing Excel file")
        products = []

        try:
            workbook = openpyxl.load_workbook(BytesIO(excel_data), data_only=True)
            worksheet = workbook.active  # Get the active worksheet

            if not worksheet:
                raise ValueError("No active worksheet found in the Excel file")

            columns = self.detect_columns(worksheet)
            logger.debug(f"Detected columns: {columns}")

            for row_idx, row in enumerate(worksheet.iter_rows(), start=1):
                # Skip header and empty rows
                if len(row) != len(columns):
                    continue

                row_map = self.map_columns(row, columns)
                if not row_map["sifra"]:
                    continue

                try:
                    product_data = {
                        "product": row_map["naziv"],
                        "product_id": row_map["sifra"],
                        "brand": row_map["marka"],
                        "barcode": row_map["barkod"],
                        "category": row_map["kategorija proizvoda"],
                        "quantity": row_map["neto kolicina"],
                        "unit": row_map["jedinica mjere"],
                        "unit_price": self.parse_price(
                            row_map["cijena za jedinicu mjere"], False
                        ),
                        "price": self.parse_price(row_map["mpc"], False),
                        "special_price": self.parse_price(
                            row_map[
                                "mpc za vrijeme posebnog oblika prodaje (rasprodaja proizvoda koji izlaze iz asortimana)"
                            ],
                            False,
                        ),
                        "best_price_30": self.parse_price(
                            row_map[
                                "najniza cijena u posljednjih 30 dana prije rasprodaje"
                            ],
                            False,
                        ),
                        "anchor_price": self.parse_price(
                            row_map[
                                "sidrena cijena na 2.5.2025. ili na datum ulistanja"
                            ],
                            False,
                        ),
                    }

                    # Apply common fixups from base class
                    product_data = self.fix_product_data(product_data)

                    # Create Product object
                    product = Product(**product_data)  # type: ignore
                    products.append(product)
                except Exception as e:
                    row_txt = "; ".join([str(cell.value or "") for cell in row])
                    logger.warning(f"Failed to parse row {row_idx}: `{row_txt}`: {e}")
                    continue

        except Exception as e:
            logger.error(f"Error parsing Excel file: {e}", exc_info=True)
            raise

        logger.debug(f"Parsed {len(products)} products from Excel file")
        return products

    def get_all_products(self, date: datetime.date) -> list[Store]:
        """
        Main method to fetch and parse all product and price info.

        Args:
            date: The date to search for in the price list.

        Returns:
            List with a single Store object containing all products.

        Raises:
            ValueError: If no price list is found for the given date.
        """
        content = self.fetch_text(self.INDEX_URL)
        if not content:
            logger.warning(f"No content found at {self.INDEX_URL}")
            return []

        # Find Excel file URL for the exact target date from JSON
        excel_url = self.find_excel_url(content, date)
        logger.info(f"Found Excel file URL: {excel_url}")

        # Download Excel file
        with TemporaryFile(mode="w+b") as temp_file:
            self.fetch_binary(excel_url, temp_file)
            temp_file.seek(0)
            excel_data = temp_file.read()

        # Parse Excel file
        products = self.parse_excel(excel_data)

        if not products:
            logger.warning(f"No products found for date {date}")
            return []

        # Create a global store
        store = Store(
            chain=self.CHAIN,
            store_type="store",
            store_id=self.STORE_ID,
            name=self.STORE_NAME,
            street_address="",
            zipcode="",
            city="",
            items=products,
        )

        return [store]


if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG)
    crawler = DmCrawler()
    stores = crawler.crawl(datetime.date.today())
    print(stores[0])
    print(stores[0].items[0])

# ============================================================================
# Eurospin CRAWLER CLASSES
# ============================================================================


logger = logging.getLogger(__name__)


class EurospinCrawler(BaseCrawler):
    """Crawler for Eurospin store prices."""

    CHAIN = "eurospin"
    BASE_URL = "https://www.eurospin.hr"
    INDEX_URL = f"{BASE_URL}/cjenik/"

    # Mapping for price fields
    PRICE_MAP = {
        # field: (column, is_required)
        "price": ("MALOPROD.CIJENA(EUR)", True),
        "unit_price": ("CIJENA_ZA_JEDINICU_MJERE", True),
        "special_price": ("MPC_POSEB.OBLIK_PROD", False),
        "best_price_30": ("NAJNIŽA_MPC_U_30DANA", False),
        "anchor_price": ("SIDRENA_CIJENA", False),
    }

    # Mapping for other fields
    FIELD_MAP = {
        "product": ("NAZIV_PROIZVODA", True),
        "product_id": ("ŠIFRA_PROIZVODA", True),
        "brand": ("MARKA_PROIZVODA", False),
        "quantity": ("NETO_KOLIČINA", False),
        "unit": ("JEDINICA_MJERE", False),
        "barcode": ("BARKOD", False),
        "category": ("KATEGORIJA_PROIZVODA", False),
    }

    STORE_ID_MAP = {
        "Ulica hrvatskog preporoda 70 Dugo Selo": "310032",
        "Ulica Rimske centurijacije 100": "310013",
        "Ulica Juraja Dobrile 1C": "310006",
        "Zagrebacka ul 49G": "310012",
        "Gacka ulica 70": "310017",
        "Ulica Istarskih narodnjaka 17 Stop Shop": "310027",
        "Zagrebacka cesta 162A": "310018",
        "Ulica Ote Horvata 1 33000 Virovitica": "310036",
        "Cesta Dalmatinskih brigada 7a": "310030",
        "Celine 2": "310009",
        "Ulica Mate Vlašica 51A": "310010",
        "Koprivnicka ulica 34A": "310033",
        "Ulica Furicevo 20": "310016",
        "Zvonarska ulica 63": "310035",
        "Ulica Petra Svacica 2B": "310014",
        "Zagrebacka 52": "310004",
        "Ulica Matije Gupca 59": "310021",
        "Ulica Mihovila P Miškine 5": "310024",
        "4 Gardijske Brigade 1": "310003",
        "Ulica hrvatskih branitelja 2": "310005",
        "Ulica Ante Starcevica 20": "310019",
        "I Štefanovecki zavoj 12": "310002",
        "Štrmac 303": "310026",
        "Ljudevita Šestica 7": "310037",
        "Ulica Vlahe Paljetka 7": "310011",
        "Ulica Veceslava Holjevca 15": "310034",
        "Stop shop": "310028",
        "Solinska ulica 84": "310015",
        "Obrtnicka ulica 2": "310008",
        "Ulica kralja Tomislava 47A": "310007",
        "Žutska ulica broj 1": "310023",
    }

    def parse_index(self, content: str) -> list[str]:
        """
        Parse the Eurospin index page to extract ZIP links.

        Args:
            content: HTML content of the index page

        Returns:
            List of ZIP urls on the page
        """
        soup = BeautifulSoup(content, "html.parser")
        urls = []

        csv_options = soup.select("option[value$='.zip']")
        for option in csv_options:
            href = str(option.get("value"))
            if href.startswith(("http://", "https://")):
                urls.append(href)
            else:
                urls.append(f"{self.BASE_URL}{href}")

        return list(set(urls))

    def parse_store_info(self, url: str) -> Store:
        """
        Extracts store information from a CSV download URL.

        Example filename:
            supermarket-Zvonarska_ulica_63-Vinkovci-32100-23.05.2025-7.30.csv:
        https://www.eurospin.hr/wp-content/themes/eurospin/documenti-prezzi/supermarket-310037-Ljudevita_Šestica_7-Karlovac-123456-21.05.2025-7.30.csv

        Args:
            url: CSV download URL with store information in the filename

        Returns:
            Store object with parsed store information
        """
        logger.debug(f"Parsing store information from URL: {url}")

        filename = os.path.basename(url)
        parts = filename.split("-")

        if len(parts) < 6:
            raise ValueError(f"Invalid CSV filename format: {filename}")

        if len(parts) == 6:
            addr = parts[1].replace("_", " ")
            store_id = self.STORE_ID_MAP.get(addr, addr)
            logger.debug(
                f"Store ID missing, assuming '{store_id}' based on address '{addr}'"
            )
            parts.insert(1, store_id)

        store_type = parts[0].lower()
        store_id = parts[1]
        street_address = parts[2].replace("_", " ")
        city = parts[3]

        # Valid zipcode is 5 digits
        zipcode = parts[4] if len(parts[4]) == 5 and parts[4].isdigit() else ""

        store = Store(
            chain=self.CHAIN,
            store_type=store_type,
            store_id=store_id,
            name=f"{self.CHAIN.capitalize()} {city}",
            street_address=street_address,
            zipcode=zipcode,
            city=city,
            items=[],
        )

        logger.info(
            f"Parsed store: {store.store_type}, {store.street_address}, {store.zipcode}, {store.city}"
        )
        return store

    def get_store_prices(self, content: bytes) -> List[Product]:
        """
        Fetch and parse store prices from a CSV URL.

        Args:
            csv_url: URL to the CSV file containing prices

        Returns:
            List of Product objects
        """
        try:
            return self.parse_csv(content.decode("windows-1250"), delimiter=";")
        except Exception as e:
            logger.error(f"Failed to get store prices: {e}", exc_info=True)
            return []

    def get_index(self, date: datetime.date) -> str | None:
        """
        Fetch and parse the index page to get ZIP URL for the specified date.

        Args:
            date: The date to search for in the price list.

        Returns:
            URL to the zip file containing CSVs with prices, or None if not found.
        """
        content = self.fetch_text(self.INDEX_URL)

        if not content:
            logger.warning(f"No content found at {self.INDEX_URL}")
            return None

        all_urls = self.parse_index(content)
        date_str = f"{date.day:02d}.{date.month:02d}.{date.year}"

        for url in all_urls:
            filename = os.path.basename(url)
            if date_str in filename:
                return url
        else:
            logger.warning(f"No URLs found matching date {date_str}")
            return None

    def get_all_products(self, date: datetime.date) -> list[Store]:
        """
        Main method to fetch and parse all store, product and price info.

        Args:
            date: The date to search for in the price list.

        Returns:
            List of Store objects with their products.

        Raises:
            ValueError: If no price list is found for the given date.
        """
        zip_url = self.get_index(date)

        if not zip_url:
            logger.warning(f"ZIP archive URL not found for date {date}")
            return []

        stores = []

        for filename, content in self.get_zip_contents(zip_url, ".csv"):
            try:
                store = self.parse_store_info(filename)
                products = self.get_store_prices(content)
            except Exception as e:
                logger.error(
                    f"Error processing store from {filename}: {e}", exc_info=True
                )
                continue

            if not products:
                logger.warning(f"No products found in {filename}, skipping")
                continue

            store.items = products
            stores.append(store)

        return stores


if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG)
    crawler = EurospinCrawler()
    stores = crawler.crawl(datetime.date.today())
    print(stores[0])
    print(stores[0].items[0])

# ============================================================================
# Kaufland CRAWLER CLASSES
# ============================================================================


logger = logging.getLogger(__name__)


class KauflandCrawler(BaseCrawler):
    """Crawler for Kaufland store prices."""

    CHAIN = "kaufland"
    BASE_URL = "https://www.kaufland.hr"
    INDEX_URL = f"{BASE_URL}/akcije-novosti/popis-mpc.html"

    # Mapping for price fields
    PRICE_MAP = {
        # field: (column, is_required)
        "price": ("maloprod.cijena(EUR)", False),
        "unit_price": ("cijena jed.mj.(EUR)", False),
        "special_price": ("MPC poseb.oblik prod", False),
        "best_price_30": ("Najniža MPC u 30dana", False),
        "anchor_price": ("Sidrena cijena", False),
    }

    # Mapping for other fields
    FIELD_MAP = {
        "product": ("naziv proizvoda", True),
        "product_id": ("šifra proizvoda", True),
        "brand": ("marka proizvoda", False),
        "quantity": ("neto količina(KG)", False),
        "unit": ("jedinica mjere", False),
        "barcode": ("barkod", False),
        "category": ("Kategorija", False),
        "anchor_date": ("Datum sidrenja", False),
    }

    CITIES = [
        "Zagreb Blato",
        "Zagreb",
        "Karlovac",
        "Velika Gorica",
        "Zapresic",
        "Zadar",
        "Cakovec",
        "Đakovo",
        "Sisak",
        "Koprivnica",
        "Slavonski Brod",
        "Nova Gradiska",
        "Sinj",
        "Rovinj",
        "Osijek",
        "Virovitica",
        "Biograd",
        "Dugo Selo",
        "Sibenik",
        "Pula",
        "Porec",
        "Makarska",
        "Kutina",
        "Split",
        "Vinkovci",
        "Rijeka",
        "Bjelovar",
        "Ivanec",
        "Trogir",
        "Umag",
        "Vukovar",
        "Zabok",
        "Cibaca",
        "Pozega",
        "Dakovo",
        "Vodice",
        "Varazdin",
        "Samobor",
    ]

    # Pattern to extract date and price from anchor price string
    # Example format: "MPC 2.5.2025=7,99€"
    ANCHOR_PRICE_PATTERN = re.compile(r"MPC\s+(\d+\.\d+\.\d+)=(.+)")

    # Pattern to parse store information from filename
    # Format: Supermarket_Put_Gaceleza_1D_Vodice_6730_15_05_2025_7_30.csv
    ADDRESS_PATTERN = re.compile(r"(Supermarket|Hipermarket)_(.+?)_(\d{4})_")

    def get_index(self, date: datetime.date) -> dict[str, str]:
        """
        Get all CSV links from the Kaufland index page.

        Args:
            date: Date to get prices for

        Returns:
            Dictionary with title → URL mappings for CSV files.
        """

        # 0. Fetch the Kaufland index page

        content = self.fetch_text(self.INDEX_URL)
        if not content:
            raise ValueError("Failed to fetch Kaufland index page")

        soup = BeautifulSoup(content, "html.parser")

        # 1. Locate the Vue AssetList component
        list_el = soup.select_one("div[data-component=AssetList]")
        if not list_el:
            raise ValueError("Failed to find CSV links in Kaufland index page")

        # 2. Extract the AssetList component settings from a prop attrib
        vue_props = loads(str(list_el.get("data-props")))

        json_url = self.BASE_URL + vue_props.get("settings", {}).get("dataUrlAssets")
        if not json_url:
            raise ValueError("Failed to find JSON URL in Kaufland index page")

        # 3. Fetch the JSON data from the URL
        logger.debug(f"Fetching JSON data from {json_url}")
        json_content = self.fetch_text(json_url)
        if not json_content:
            raise ValueError("Failed to fetch JSON data from Kaufland index page")

        # 4. Parse the JSON data to extract CSV URLs
        json_data = loads(json_content)

        urls = {}
        date_str = date.strftime("_%d_%m_%Y_")
        date_str2 = date.strftime("_%d%m%Y_")
        for item in json_data:
            label = item.get("label")
            url = item.get("path")
            if not label or not url:
                continue
            if date_str not in label and date_str2 not in label:
                continue
            urls[label] = f"{self.BASE_URL}{url}"

        return urls

    def parse_store_info(self, title: str) -> Store:
        """
        Extract store information from the CSV title.

        Args:
            title: Title of the CSV file

        Returns:
            Store object with parsed information
        """
        # Format example: Supermarket_Put_Gaceleza_1D_Vodice_6730_15_05_2025_7_30.csv
        match = self.ADDRESS_PATTERN.search(title)
        if not match:
            raise ValueError(f"Could not parse store info from filename: {title}")

        store_type, address_part, store_id = match.groups()

        store_type = store_type.lower()
        street_address = address_part.replace("_", " ").title()
        city = ""

        # Look for cities in the address
        for city_name in self.CITIES:
            if self.strip_diacritics(street_address).endswith(city_name):
                city = city_name
                street_address = street_address[: -len(city_name)].strip()
                break

        # Create store object
        store = Store(
            chain=self.CHAIN,
            store_type=store_type,
            store_id=store_id,
            name=f"{self.CHAIN.capitalize()} {city}",
            street_address=street_address,
            city=city,
            zipcode="",
            items=[],
        )

        logger.info(
            f"Parsed store: {store.store_type} ({store.store_id}), {store.street_address}, {store.city}"
        )
        return store

    def get_store_prices(self, csv_url: str) -> List[Product]:
        """
        Get and parse prices from a store's CSV file.

        Args:
            csv_url: URL of the CSV file

        Returns:
            List of Product objects
        """
        try:
            content = self.fetch_text(csv_url, encodings=["windows-1250"])
            return self.parse_csv(content, delimiter="\t")
        except Exception as e:
            logger.error(
                f"Failed to get store prices from {csv_url}: {e}",
                exc_info=True,
            )
            return []

    def parse_csv_row(self, row: dict) -> Product:
        anchor_price = row.get("Sidrena cijena")
        row["Datum sidrenja"] = ""

        if anchor_price:
            match = self.ANCHOR_PRICE_PATTERN.search(anchor_price)
            if match:
                date_str, price_str = match.groups()

                try:
                    row["Datum sidrenja"] = (
                        datetime.datetime.strptime(
                            date_str,
                            "%d.%m.%Y",
                        )
                        .date()
                        .strftime("%Y-%m-%d")
                    )
                    row["Sidrena cijena"] = price_str
                except (ValueError, IndexError) as e:
                    logger.warning(f"Error parsing anchor price {anchor_price}: {e}")
                    row["Sidrena cijena"] = ""
            else:
                row["Sidrena cijena"] = ""

        return super().parse_csv_row(row)

    def get_all_products(self, date: datetime.date) -> list[Store]:
        """
        Main method to fetch and parse all store, product and price info.

        Args:
            date: The date to search for in the price list.

        Returns:
            List of Store objects with their products.
        """
        csv_links = self.get_index(date)
        stores = []

        for title, url in csv_links.items():
            try:
                store = self.parse_store_info(title)
                products = self.get_store_prices(url)
            except Exception as e:
                logger.error(f"Error processing store from {url}: {e}", exc_info=True)
                continue

            if not products:
                logger.warning(f"No products found for {url}, skipping")
                continue

            store.items = products
            stores.append(store)

        return stores


if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG)
    crawler = KauflandCrawler()
    stores = crawler.crawl(datetime.date.today() - datetime.timedelta(days=1))
    print(stores[0])
    print(stores[0].items[0])

# ============================================================================
# Konzum CRAWLER CLASSES
# ============================================================================

 
logger = logging.getLogger(__name__)


class KonzumCrawler(BaseCrawler):
    """Crawler for Konzum store prices."""

    CHAIN = "konzum"
    BASE_URL = "https://www.konzum.hr"
    INDEX_URL = f"{BASE_URL}/cjenici"

    # Mapping for price fields
    PRICE_MAP = {
        # field: (column, is_required)
        "price": ("MALOPRODAJNA CIJENA", False),
        "unit_price": ("CIJENA ZA JEDINICU MJERE", True),
        "special_price": ("MPC ZA VRIJEME POSEBNOG OBLIKA PRODAJE", False),
        "best_price_30": ("NAJNIŽA CIJENA U POSLJEDNJIH 30 DANA", False),
        "anchor_price": ("SIDRENA CIJENA NA 2.5.2025", False),
    }

    # Mapping for other fields
    FIELD_MAP = {
        "product": ("NAZIV PROIZVODA", True),
        "product_id": ("ŠIFRA PROIZVODA", True),
        "brand": ("MARKA PROIZVODA", False),
        "quantity": ("NETO KOLIČINA", False),
        "unit": ("JEDINICA MJERE", False),
        "barcode": ("BARKOD", False),
        "category": ("KATEGORIJA PROIZVODA", False),
    }

    ADDRESS_PATTERN = re.compile(r"(.*) (\d{5}) (.*)")

    def parse_index(self, content: str) -> list[str]:
        """
        Parse the Konzum index page to extract the price date and CSV links.

        Args:
            content: HTML content of the index page

        Returns:
            List of CSV urls on the page
        """

        soup = BeautifulSoup(content, "html.parser")

        urls = []
        csv_links = soup.select("a[format='csv']")

        for link in csv_links:
            href = link.get("href")
            if href:
                urls.append(f"{self.BASE_URL}{href}")

        return list(set(urls))

    def parse_store_info(self, url: str) -> Store:
        """
        Extracts store information from a CSV download URL.

        Args:
            url: CSV download URL with store information in the query parameters

        Returns:
            Store object with parsed store information, or None if parsing fails
        """

        logger.debug(f"Parsing store information from URL: {url}")

        parsed_url = urllib.parse.urlparse(url)
        query_params = urllib.parse.parse_qs(parsed_url.query)
        title = urllib.parse.unquote(query_params.get("title", [""])[0])
        title = title.replace("_", " ")

        if not title:
            raise ValueError(f"No title parameter found in URL: {url}")

        logger.debug(f"Decoded title: {title}")

        parts = [part.strip() for part in title.split(",")]
        if len(parts) < 6:  # Ensure we have the expected number of parts
            raise ValueError(f"Invalid CSV title format: {title}")

        # Extract store type
        store_type = (parts[0]).lower()
        store_id = parts[2] if len(parts) == 6 else parts[3]

        # Format:
        # SUPERMARKET,REPUBLIKE 1 31300 BELI MANASTIR,0904,1629,21.05.2025, 05-22.CSV
        # SUPERMARKET,CARLOTTA GRISI 5, SVETI ANTON 52466 NOVIGRAD,3274,1332,19.05.2025, 05-52.CSV
        m = self.ADDRESS_PATTERN.match(
            parts[1] if len(parts) == 6 else f"{parts[1]} {parts[2]}"
        )
        if not m:
            raise ValueError(f"Could not parse address from: {parts[1]}")

        # Extract address components
        street_address = m.group(1).strip().title()
        zipcode = m.group(2).strip()
        city = m.group(3).strip().title()

        store = Store(
            chain=self.CHAIN,
            store_type=store_type,
            store_id=store_id,
            name=f"{self.CHAIN.capitalize()} {city}",
            street_address=street_address,
            zipcode=zipcode,
            city=city,
            items=[],
        )

        logger.info(
            f"Parsed store: {store.store_type}, {store.street_address}, {store.zipcode}, {store.city}"
        )
        return store

    def get_index(self, date: datetime.date) -> list[str]:
        url = f"{self.INDEX_URL}?date={date:%Y-%m-%d}"

        csv_urls = []
        for page in range(1, 10):
            page_url = f"{url}&page={page}"
            content = self.fetch_text(page_url)
            if not content:
                break

            csv_urls_on_page = self.parse_index(content)
            if not csv_urls_on_page:
                break

            csv_urls.extend(csv_urls_on_page)

        return csv_urls

    def get_store_prices(self, csv_url: str) -> List[Product]:
        try:
            content = self.fetch_text(csv_url)
            return self.parse_csv(content)
        except Exception as e:
            logger.error(
                f"Failed to get store prices from {csv_url}: {e}",
                exc_info=True,
            )
            return []

    def get_all_products(self, date: datetime.date) -> list[Store]:
        """
        Main method to fetch and parse all store, product and price info.

        Args:
            date: The date to search for in the price list.

        Returns:
            List of Store objects with their products.

        Raises:
            ValueError: If no price list is found for the given date.
        """

        csv_links = self.get_index(date)
        stores = []

        for url in csv_links:
            try:
                store = self.parse_store_info(url)
                products = self.get_store_prices(url)
            except Exception as e:
                logger.error(f"Error processing store from {url}: {e}", exc_info=True)
                continue

            if not products:
                logger.warning(f"Error getting prices from {url}, skipping")
                continue
            store.items = products
            stores.append(store)

        return stores


if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG)
    crawler = KonzumCrawler()
    stores = crawler.crawl(datetime.date.today())
    print(stores[0])
    print(stores[0].items[0])

# ============================================================================
# Ktc CRAWLER CLASSES
# ============================================================================


logger = logging.getLogger(__name__)


class KtcCrawler(BaseCrawler):
    """Crawler for KTC store prices."""

    CHAIN = "ktc"
    BASE_URL = "https://www.ktc.hr"
    INDEX_URL = f"{BASE_URL}/cjenici"

    # CSV fields mapping
    PRICE_MAP = {
        # field: (column, is_required)
        "price": ("Maloprodajna cijena", True),
        "unit_price": ("Cijena za jedinicu mjere", True),
        "special_price": ("MPC za vrijeme posebnog oblika prodaje", False),
        "best_price_30": ("Najniža cijena u posljednjih 30 dana", False),
        "anchor_price": ("Sidrena cijena na 2.5.2025", False),
    }

    # Mapping for other fields
    FIELD_MAP = {
        "product": ("Naziv proizvoda", True),
        "product_id": ("Šifra proizvoda", True),
        "brand": ("Marka proizvoda", False),
        "quantity": ("Neto količina", False),
        "unit": ("Jedinica mjere", False),
        "barcode": ("Barkod", False),
        "category": ("Kategorija", False),
    }

    CITIES = [
        "KRIZEVCI",
        "VARAZDIN",
        "BJELOVAR",
        "CAKOVEC",
        "DARUVAR",
        "DUGO SELO",
        "DURDEVAC",
        "GRUBISNO POLJE",
        "IVANEC",
        "JALZABET",
        "KARLOVAC",
        "KOPRIVNICA",
        "KRAPINA",
        "KUTINA",
        "MURSKO SREDISCE",
        "PAKRAC",
        "PETRINJA",
        "PITOMACA",
        "POZEGA",
        "PRELOG",
        "SISAK II",
        "SISAK",
        "SLATINA",
        "VELIKA GORICA",
        "VIROVITICA",
        "VRBOVEC",
        "ZABOK",
        "CAZMA",
    ]

    def parse_index(self) -> list[str]:
        """
        Parse the KTC index page to extract store pages.

        Returns:
            List of store page URLs
        """
        content = self.fetch_text(self.INDEX_URL)
        soup = BeautifulSoup(content, "html.parser")

        store_urls = []
        store_links = soup.select('a[href^="cjenici?poslovnica="]')

        for link in store_links:
            href = link.get("href")
            if href:
                store_urls.append(f"{self.BASE_URL}/{href}")

        return list(set(store_urls))

    def get_store_csv_url(self, store_url: str, date: datetime.date) -> str:
        """
        Fetch the store page and extract the CSV URL for the specified date.

        Args:
            store_url: URL to the store's price list page
            date: The date to search for in the CSV filename

        Returns:
            CSV URL for the specified store and date, or None if not found
        """
        content = self.fetch_text(store_url)
        soup = BeautifulSoup(content, "html.parser")

        date_str = date.strftime("%Y%m%d")
        csv_links = soup.select('a[href$=".csv"]')

        for link in csv_links:
            href = str(link.get("href"))
            if date_str in href:
                if href.startswith("/"):
                    return f"{self.BASE_URL}{href}"
                else:
                    return f"{self.BASE_URL}/{href}"

        raise ValueError(f"No CSV found for date {date} at {store_url}")

    def parse_store_info(self, csv_url: str) -> Store:
        """
        Extracts store information from a CSV download URL.


        Format example (URL path basename):
            `TRGOVINA-SENJSKA ULICA 118 KARLOVAC-PJ8A-1-20250515-071626.csv`

        Args:
            csv_url: CSV download URL with store information

        Returns:
            Store object with parsed store information
        """
        logger.debug(f"Parsing store information from URL: {csv_url}")

        parsed_url = urlparse(csv_url)
        path_parts = parsed_url.path.split("/")

        # Get the last two parts of the path
        csv_filename = unquote(path_parts[-1])

        # Try to guess the city if possible
        for city in self.CITIES:
            if city in csv_filename:
                break
        else:
            city = ""

        # Parse csv_filename to get store type and address
        parts = csv_filename.split("-")
        if len(parts) < 3:
            raise ValueError(f"Invalid CSV filename format: {csv_filename}")

        store_type = parts[0].lower()
        store_id = parts[2]

        # Address is the second part, but might contain the city name too
        street_address = parts[1].strip()
        if city:
            # Remove the city name to get just the street address
            street_address = street_address.replace(city, "").strip()
            street_address = re.sub(r"\s+", " ", street_address)

        # Create the store object
        store = Store(
            chain=self.CHAIN,
            store_type=store_type,
            store_id=f"PJ{store_id}",
            name=f"{self.CHAIN.upper()} {city}",
            street_address=street_address.title(),
            zipcode="",  # No ZIP code in the URL
            city=city.title(),
            items=[],
        )

        logger.info(
            f"Parsed store: {store.store_type}, {store.street_address}, {store.city}"
        )
        return store

    def get_store_prices(self, csv_url: str) -> List[Product]:
        """
        Fetch and parse CSV content to extract product prices.

        Args:
            csv_url: URL of the CSV file to download

        Returns:
            List of Product objects
        """
        try:
            # KTC CSVs are encoded in Windows-1250
            content = self.fetch_text(csv_url, encodings=["windows-1250"])
            return self.parse_csv(content, delimiter=";")
        except Exception as e:
            logger.error(
                f"Failed to get store prices from {csv_url}: {e}",
                exc_info=True,
            )
            return []

    def get_all_products(self, date: datetime.date) -> list[Store]:
        """
        Main method to fetch and parse all store, product and price info.

        Args:
            date: The date to search for in the price list.

        Returns:
            List of Store objects with their products.

        Raises:
            ValueError: If no price list is found for the given date.
        """
        store_urls = self.parse_index()
        stores = []

        for store_url in store_urls:
            try:
                csv_url = self.get_store_csv_url(store_url, date)
                if not csv_url:
                    logger.warning(f"No CSV found for date {date} at {store_url}")
                    continue

                store = self.parse_store_info(csv_url)
                products = self.get_store_prices(csv_url)

                if not products:
                    logger.warning(f"No products found in {csv_url}, skipping")
                    continue

                store.items = products
                stores.append(store)

            except Exception as e:
                logger.error(
                    f"Error processing store from {store_url}: {e}", exc_info=True
                )
                continue

        return stores


if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG)
    crawler = KtcCrawler()
    stores = crawler.crawl(datetime.date.today())
    print(stores[0])
    print(stores[0].items[0])


# ============================================================================
# Lidl CRAWLER CLASSES
# ============================================================================


logger = logging.getLogger(__name__)


class LidlCrawler(BaseCrawler):
    """
    Crawler for Lidl store prices.

    This class handles downloading and parsing price data from Lidl's website.
    It fetches the price list index page, finds the ZIP for the specified date,
    downloads and extracts it, and parses the CSV files inside.
    """

    CHAIN = "lidl"
    BASE_URL = "https://tvrtka.lidl.hr"
    INDEX_URL = f"{BASE_URL}/cijene"
    TIMEOUT = 180.0  # Longer timeout for ZIP download
    ZIP_DATE_PATTERN = re.compile(
        r".*/Popis_cijena_po_trgovinama_na_dan_(\d{1,2})_(\d{1,2})_(\d{4})\.zip"
    )

    ANCHOR_PRICE_COLUMN = "Sidrena_cijena_na_02.05.2025"
    PRICE_MAP = {
        "price": ("MALOPRODAJNA_CIJENA", False),
        "unit_price": ("CIJENA_ZA_JEDINICU_MJERE", False),
        "special_price": ("MPC_ZA_VRIJEME_POSEBNOG_OBLIKA_PRODAJE", False),
        "anchor_price": (ANCHOR_PRICE_COLUMN, False),
    }

    FIELD_MAP = {
        "product": ("NAZIV", False),
        "product_id": ("ŠIFRA", True),
        "brand": ("MARKA", False),
        "quantity": ("NETO_KOLIČINA", False),
        "unit": ("JEDINICA_MJERE", False),
        "barcode": ("BARKOD", False),
        "category": ("KATEGORIJA_PROIZVODA", False),
        "packaging": ("PAKIRANJE", False),
    }

    ADDRESS_PATTERN = re.compile(
        r"^(Supermarket)\s+"  # 'Supermarket'
        r"(\d+)_+"  # store number (digits)
        r"([\w._\s-]+?)_+"  # address (lazy match, allows spaces, underscores, dots)
        r"(\d{5})_+"  # ZIP code (5 digits)
        r"([A-ZŠĐČĆŽ_\s-]+?)_"  # city (letters, underscores or spaces, lazy match)
        r".*\.csv",  # the rest
        re.UNICODE | re.IGNORECASE,
    )

    def parse_store_from_filename(self, filename: str) -> Optional[Store]:
        """
        Extract store information from CSV filename using filename parts.

        Args:
            filename: Name of the CSV file with store information

        Returns:
            Store object with parsed store information, or None if parsing fails
        """
        logger.debug(f"Parsing store information from filename: {filename}")

        try:
            m = self.ADDRESS_PATTERN.match(filename)
            if not m:
                logger.warning(f"Filename doesn't match expected pattern: {filename}")
                return None

            store_type, store_id, address, zipcode, city = m.groups()
            city = city.replace("_", " ")
            address = address.replace("_", " ")
            if address.startswith(city + " "):
                address = address[len(city) + 1 :]
                if address.startswith("-"):
                    address = address[1:]

            store = Store(
                chain=self.CHAIN,
                store_id=store_id,
                name=f"Lidl {city}",
                store_type=store_type.lower(),
                city=city.title(),
                street_address=address.strip().title(),
                zipcode=zipcode,
                items=[],
            )

            logger.info(
                f"Parsed store: {store.name}, {store.store_type}, {store.city}, {store.street_address}, {store.zipcode}"
            )
            return store

        except Exception as e:
            logger.error(f"Failed to parse store from filename {filename}: {str(e)}")
            return None

    def parse_csv_row(self, row: dict) -> Product:
        anchor_price = row.get(self.ANCHOR_PRICE_COLUMN, "").strip()
        if "Nije_bilo_u_prodaji" in anchor_price:
            row[self.ANCHOR_PRICE_COLUMN] = None

        return super().parse_csv_row(row)

    def get_index(self, date: datetime.date) -> str:
        content = self.fetch_text(self.INDEX_URL)
        zip_urls_by_date = self.parse_index_for_zip(content)
        others = ", ".join(f"{d:%Y-%m-%d}" for d in zip_urls_by_date)
        logger.debug(f"Available price lists: {others}")
        if date not in zip_urls_by_date:
            raise ValueError(f"No price list found for {date}")
        return zip_urls_by_date[date]

    def get_all_products(self, date: datetime.date) -> list[Store]:
        """
        Main method to fetch and parse all products from Lidl's price lists.

        Args:
            date: The date for which to fetch the price list

        Returns:
            Tuple with the date and the list of Store objects,
            each containing its products.

        Raises:
            ValueError: If the price list ZIP cannot be found or processed
        """
        zip_url = self.get_index(date)
        stores = []

        for filename, content in self.get_zip_contents(zip_url, ".csv"):
            logger.debug(f"Processing file: {filename}")
            store = self.parse_store_from_filename(filename)
            if not store:
                logger.warning(f"Skipping CSV {filename} due to store parsing failure")
                continue

            # Parse CSV and add products to the store
            text = content.decode("windows-1250")
            headers = text.splitlines()[0]
            if "\t" in headers:
                delimiter = "\t"
            elif ";" in headers:
                delimiter = ";"
            elif "," in headers:
                delimiter = ","
            else:
                logger.warning(f"Unknown delimiter in CSV: {filename}; ignoring")
                continue
            products = self.parse_csv(text, delimiter=delimiter)
            store.items = products
            stores.append(store)

        return stores


if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG)
    crawler = LidlCrawler()
    stores = crawler.get_all_products(datetime.date(2025, 5, 17))
    print(stores[0])
    print(stores[0].items[0])

# ============================================================================
# Metro CRAWLER CLASSES
# ============================================================================


logger = logging.getLogger(__name__)


class MetroCrawler(BaseCrawler):
    """Crawler for Metro store prices."""

    CHAIN = "metro"
    BASE_URL = "https://metrocjenik.com.hr"

    # Regex to parse store information from the filename
    # Format: <store_type>_METRO_YYYYMMDDTHHMM_<store_id>_<address>,<city>.csv
    # Example: skladiste_za_trgovanje_robom_na_veliko_i_malo_METRO_20250521T1149_S20_CESTA_PAPE_IVANA_PAVLA_II_3,_KASTEL_SUCURAC.csv
    STORE_FILENAME_PATTERN = re.compile(
        r"^(?P<store_type>.+?)_METRO_\d{8}T\d{4}_"
        r"(?P<store_id>[^_]+)_"
        r"(?P<address>[^,]+),"
        r"(?P<city>[^.]+)\.csv$"
    )

    # Mapping for price fields from CSV columns
    PRICE_MAP = {
        # field: (column_name, is_required)
        "price": ("MPC", True),
        "unit_price": ("CIJENA_PO_MJERI", True),
        "special_price": ("POSEBNA_PRODAJA", False),
        "best_price_30": ("NAJNIZA_30_DANA", False),
        "anchor_price": ("SIDRENA_02_05", False),
    }

    # Mapping for other product fields from CSV columns
    FIELD_MAP = {
        "product": ("NAZIV", True),
        "product_id": ("SIFRA", True),
        "brand": ("MARKA", False),
        "quantity": ("NETO_KOLICINA", False),
        "unit": ("JED_MJERE", False),
        "barcode": ("BARKOD", False),
        "category": ("KATEGORIJA", False),
    }

    def parse_index(self, content: str) -> list[str]:
        """
        Parse the Metro index page to extract CSV links.

        Args:
            content: HTML content of the index page

        Returns:
            List of absolute CSV URLs on the page
        """
        soup = BeautifulSoup(content, "html.parser")
        urls = []

        for link_tag in soup.select('a[href$=".csv"]'):
            href = str(link_tag.get("href"))
            if href:
                full_url = f"{self.BASE_URL}/{href.lstrip('/')}"
                urls.append(full_url)

        return list(set(urls))  # Return unique URLs

    def parse_store_info(self, url: str) -> Store:
        """
        Extracts store information from a CSV download URL.

        Example URL path part:
        skladiste_za_trgovanje_robom_na_veliko_i_malo_METRO_20250521T1149_S20_CESTA_PAPE_IVANA_PAVLA_II_3%2C_KASTEL_SUCURAC.csv

        Args:
            url: CSV download URL with store information in the filename

        Returns:
            Store object with parsed store information
        """
        logger.debug(f"Parsing store information from Metro URL: {url}")

        filename = unquote(os.path.basename(url))

        match = self.STORE_FILENAME_PATTERN.match(filename)
        if not match:
            raise ValueError(f"Invalid CSV filename format for Metro: {filename}")

        data = match.groupdict()

        store_type = data["store_type"].replace("_", " ").lower()
        store_id = data["store_id"]
        # Address: "CESTA_PAPE_IVANA_PAVLA_II_3" -> "Cesta Pape Ivana Pavla Ii 3"
        address_raw = data["address"]
        street_address = address_raw.replace("_", " ").title()
        # City: "_KASTEL_SUCURAC" -> "Kastel Sucurac" (strip potential leading/trailing _ from regex capture)
        city_raw = data["city"]
        city = city_raw.strip("_").replace("_", " ").title()

        store = Store(
            chain=self.CHAIN,
            store_type=store_type,
            store_id=store_id,
            name=f"{self.CHAIN.capitalize()} {city} {store_id}",  # e.g. "Metro Kastel Sucurac S20"
            street_address=street_address,
            zipcode="",  # Zipcode is not available in the filename
            city=city,
            items=[],
        )

        logger.info(
            f"Parsed Metro store: {store.name}, Type: {store.store_type}, Address: {store.street_address}, City: {store.city}"
        )
        return store

    def get_store_prices(self, csv_url: str) -> List[Product]:
        """
        Fetch and parse store prices from a Metro CSV URL.
        The CSV is comma-separated and UTF-8 encoded.

        Args:
            csv_url: URL to the CSV file containing prices

        Returns:
            List of Product objects
        """
        try:
            # fetch_text handles potential HTTP errors. CSV is UTF-8 by default from response.text.
            content = self.fetch_text(csv_url)
            # Metro CSVs are comma-delimited
            return self.parse_csv(content, delimiter=",")
        except Exception as e:
            logger.error(
                f"Failed to get Metro store prices from {csv_url}: {e}",
                exc_info=True,
            )
            return []

    def get_index(self, date: datetime.date) -> list[str]:
        """
        Fetch and parse the Metro index page to get CSV URLs for the specified date.

        Args:
            date: The date to search for in the price list (YYYYMMDD format).

        Returns:
            List of CSV URLs containing prices for the specified date.
        """
        content = self.fetch_text(self.BASE_URL)

        if not content:
            logger.warning(f"No content found at Metro index URL: {self.BASE_URL}")
            return []

        all_urls = self.parse_index(content)
        # Date format in Metro filenames is YYYYMMDD, e.g., _METRO_20250521T...
        date_str = date.strftime("%Y%m%d")

        matching_urls = []
        for url in all_urls:
            filename = os.path.basename(url)
            # Check if the YYYYMMDD date string (followed by 'T' for time) is in the filename
            if f"_{date_str}T" in filename:
                matching_urls.append(url)

        if not matching_urls:
            logger.warning(f"No Metro URLs found matching date {date:%Y-%m-%d}")

        return matching_urls

    def get_all_products(self, date: datetime.date) -> list[Store]:
        """
        Main method to fetch and parse all Metro store, product, and price info for a given date.

        Args:
            date: The date to search for in the price list.

        Returns:
            List of Store objects with their products.
        """
        csv_links = self.get_index(date)

        if not csv_links:
            logger.warning(f"No Metro CSV links found for date {date.isoformat()}")
            return []

        stores = []
        for url in csv_links:
            try:
                store = self.parse_store_info(url)
                products = self.get_store_prices(url)
            except ValueError as ve:  # Catch specific error from parse_store_info
                logger.error(
                    f"Skipping store due to parsing error from URL {url}: {ve}",
                    exc_info=False,
                )  # exc_info=False to reduce noise for expected parsing errors
                continue
            except Exception as e:
                logger.error(
                    f"Error processing Metro store from {url}: {e}", exc_info=True
                )
                continue  # Skip to the next URL on error

            if not products:
                logger.warning(f"No products found for Metro store at {url}, skipping.")
                continue

            store.items = products
            stores.append(store)

        return stores


if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG)
    crawler = MetroCrawler()
    stores = crawler.crawl(datetime.date.today())
    print(stores[0])
    print(stores[0].items[0])

# ============================================================================
# Ntl CRAWLER CLASSES
# ============================================================================


logger = logging.getLogger(__name__)


class NtlCrawler(BaseCrawler):
    """Crawler for NTL store prices."""

    CHAIN = "ntl"
    BASE_URL = "https://www.ntl.hr/cjenici-za-ntl-supermarkete"

    # Regex to parse store information from the filename
    # Format: Supermarket_Ljudevita Gaja 1_DUGA RESA_10103_263_25052025_07_22_36.csv
    STORE_FILENAME_PATTERN = re.compile(
        r"(?P<store_type>[^_]+)_(?P<street_address>[^_]+)_(?P<city>[^_]+)_(?P<store_id>\d+)_.*\.csv$"
    )

    # Mapping for price fields from CSV columns
    PRICE_MAP = {
        # field: (column_name, is_required)
        "price": ("Maloprodajna cijena", False),
        "unit_price": ("Cijena za jedinicu mjere", False),
        "special_price": ("MPC za vrijeme posebnog oblika prodaje", False),
        "anchor_price": ("Sidrena cijena na 2.5.2025", False),
    }

    # Mapping for other product fields from CSV columns
    FIELD_MAP = {
        "product_id": ("Šifra proizvoda", True),
        "barcode": ("Barkod", False),
        "product": ("Naziv proizvoda", True),
        "brand": ("Marka proizvoda", False),
        "quantity": ("Neto količina", False),
        "unit": ("Jedinica mjere", False),
        "category": ("Kategorija proizvoda", False),
    }

    def parse_index(self, content: str) -> list[str]:
        """
        Parse the NTL index page to extract CSV links.

        Args:
            content: HTML content of the index page

        Returns:
            List of absolute CSV URLs on the page
        """
        soup = BeautifulSoup(content, "html.parser")
        urls = []

        for link_tag in soup.select('table a[href$=".csv"]'):
            href = str(link_tag.get("href"))
            urls.append(href)

        return list(set(urls))  # Return unique URLs

    def parse_store_info(self, url: str) -> Store:
        """
        Extracts store information from a CSV download URL.

        Example URL:
        https://www.ntl.hr/csv_files/Supermarket_Ljudevita Gaja 1_DUGA RESA_10103_263_25052025_07_22_36.csv

        Args:
            url: CSV download URL with store information in the filename

        Returns:
            Store object with parsed store information
        """
        logger.debug(f"Parsing store information from NTL URL: {url}")

        filename = unquote(os.path.basename(url))

        match = self.STORE_FILENAME_PATTERN.match(filename)
        if not match:
            raise ValueError(f"Invalid CSV filename format for NTL: {filename}")

        data = match.groupdict()

        store_type = data["store_type"].lower()
        street_address = data["street_address"]
        city = data["city"].title()
        store_id = data["store_id"]

        store = Store(
            chain=self.CHAIN,
            store_type=store_type,
            store_id=store_id,
            name=f"NTL {city}",
            street_address=street_address,
            zipcode="",  # Zipcode is not available in the filename
            city=city,
            items=[],
        )

        logger.info(
            f"Parsed NTL store: {store.name}, Address: {store.street_address}, City: {store.city}"
        )
        return store

    def get_store_prices(self, csv_url: str) -> list[Product]:
        """
        Fetch and parse store prices from an NTL CSV URL.
        The CSV is semicolon-separated and windows-1250 encoded.

        Args:
            csv_url: URL to the CSV file containing prices

        Returns:
            List of Product objects
        """
        try:
            content = self.fetch_text(csv_url, encodings=["windows-1250"])
            return self.parse_csv(content, delimiter=";")
        except Exception as e:
            logger.error(
                f"Failed to get NTL store prices from {csv_url}: {e}",
                exc_info=True,
            )
            return []

    def get_index(self, date: datetime.date) -> list[str]:
        """
        Fetch and parse the NTL index page to get CSV URLs.

        Note: NTL only shows current CSV files, so the date parameter is ignored.

        Args:
            date: The date parameter (ignored for NTL)

        Returns:
            List of all CSV URLs available on the index page.
        """
        logger.warning(
            f"NTL crawler ignores date parameter ({date:%Y-%m-%d}) - "
            "only current CSV files are available"
        )

        content = self.fetch_text(self.BASE_URL)

        if not content:
            logger.warning(f"No content found at NTL index URL: {self.BASE_URL}")
            return []

        all_urls = self.parse_index(content)

        if not all_urls:
            logger.warning("No NTL CSV URLs found on index page")

        return all_urls

    def get_all_products(self, date: datetime.date) -> list[Store]:
        """
        Main method to fetch and parse all NTL store, product, and price info.

        Note: Date parameter is ignored as NTL only provides current prices.

        Args:
            date: The date parameter (ignored for NTL)

        Returns:
            List of Store objects with their products.
        """
        csv_links = self.get_index(date)

        if not csv_links:
            logger.warning("No NTL CSV links found")
            return []

        stores = []
        for url in csv_links:
            try:
                store = self.parse_store_info(url)
                products = self.get_store_prices(url)
            except ValueError as ve:
                logger.error(
                    f"Skipping store due to parsing error from URL {url}: {ve}",
                    exc_info=False,
                )
                continue
            except Exception as e:
                logger.error(
                    f"Error processing NTL store from {url}: {e}", exc_info=True
                )
                continue

            if not products:
                logger.warning(f"No products found for NTL store at {url}, skipping.")
                continue

            store.items = products
            stores.append(store)

        return stores

    def fix_product_data(self, data: dict) -> dict:
        """
        Clean and fix NTL-specific product data.

        Args:
            data: Dictionary containing the row data

        Returns:
            The cleaned data
        """
        if "product" in data and data["product"]:
            data["product"] = data["product"].strip()

        # Call parent method for common fixups
        return super().fix_product_data(data)


if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG)
    crawler = NtlCrawler()
    stores = crawler.crawl(datetime.date.today())
    print(stores[0])
    print(stores[0].items[0])


# ============================================================================
# Plodine CRAWLER CLASSES
# ============================================================================


logger = logging.getLogger(__name__)


class PlodineCrawler(BaseCrawler):
    """
    Crawler for Plodine store prices.

    This class handles downloading and parsing price data from Plodine's website.
    It fetches the price list index page, finds the ZIP for the specified date,
    downloads and extracts it, and parses the CSV files inside.
    """

    CHAIN = "plodine"
    BASE_URL = "https://www.plodine.hr"
    INDEX_URL = f"{BASE_URL}/info-o-cijenama"
    ZIP_DATE_PATTERN = re.compile(r".*/cjenici/cjenici_(\d{2})_(\d{2})_(\d{4})_.*\.zip")

    PRICE_MAP = {
        "price": ("Maloprodajna cijena", False),
        "unit_price": ("Cijena po JM", False),
        "special_price": (
            "MPC za vrijeme posebnog oblika prodaje",
            False,
        ),
        "best_price_30": ("Najniza cijena u poslj. 30 dana", False),
        "anchor_price": ("Sidrena cijena na 2.5.2025", False),
    }

    FIELD_MAP = {
        "product": ("Naziv proizvoda", True),
        "product_id": ("Sifra proizvoda", True),
        "brand": ("Marka proizvoda", False),
        "quantity": ("Neto kolicina", False),
        "unit": ("Jedinica mjere", False),
        "barcode": ("Barkod", False),
        "category": ("Kategorija proizvoda", False),
    }

    def get_index(self, date: datetime.date) -> str:
        content = self.fetch_text(self.INDEX_URL)
        zip_urls_by_date = self.parse_index_for_zip(content)
        others = ", ".join(f"{d:%Y-%m-%d}" for d in zip_urls_by_date)
        logger.debug(f"Available price lists: {others}")
        if date not in zip_urls_by_date:
            raise ValueError(f"No price list found for {date}")
        return zip_urls_by_date[date]

    def parse_store_from_filename(self, filename: str) -> Optional[Store]:
        """
        Extract store information from CSV filename using regex.

        Example filename format:
            SUPERMARKET_SJEVERNA_VEZNA_CESTA_31_35000_SLAVONSKI_BROD_022_6_20052025014212.csv
            SUPERMARKET_ULICA_FRANJE_TUDJMANA_83A_10450_JASTREBARSKO_063_2_16052025020937.csv

        Args:
            filename: Name of the CSV file with store information

        Returns:
            Store object with parsed store information, or None if parsing fails
        """
        logger.debug(f"Parsing store information from filename: {filename}")

        try:
            pattern = (
                r"^(SUPERMARKET|HIPERMARKET)_(.+?)_(\d{5})_(.+)_(\d+)_\d+_\d+.*\.csv$"
            )
            match = re.match(pattern, filename)

            if not match:
                logger.warning(f"Failed to match filename pattern: {filename}")
                return None

            store_type, street_address, zipcode, city, store_id = match.groups()

            city = city.replace("_", " ").title()

            store = Store(
                chain="plodine",
                store_id=store_id,
                name=f"Plodine {city}",
                store_type=store_type.lower(),
                city=city,
                street_address=street_address.replace("_", " ").title(),
                zipcode=zipcode,
                items=[],
            )

            logger.info(
                f"Parsed store: {store.name} ({store.store_id}), {store.store_type}, {store.city}, {store.street_address}, {store.zipcode}"
            )
            return store

        except Exception as e:
            logger.error(f"Failed to parse store from filename {filename}: {str(e)}")
            return None

    def get_all_products(self, date: datetime.date) -> list[Store]:
        """
        Main method to fetch and parse all products from Plodine's price lists.

        Args:
            date: The date for which to fetch the price list

        Returns:
            Tuple with the date and the list of Store objects,
            each containing its products.

        Raises:
            ValueError: If the price list ZIP cannot be found or processed
        """
        zip_url = self.get_index(date)
        stores = []

        for filename, content in self.get_zip_contents(zip_url, ".csv"):
            logger.debug(f"Processing file: {filename}")
            store = self.parse_store_from_filename(filename)
            if not store:
                logger.warning(f"Skipping CSV {filename} due to store parsing failure")
                continue

            # Parse CSV and add products to the store
            products = self.parse_csv(content.decode("utf-8"), delimiter=";")
            store.items = products
            stores.append(store)

        return stores


if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG)
    crawler = PlodineCrawler()
    stores = crawler.get_all_products(datetime.date.today())
    print(stores[0])
    print(stores[0].items[0])

# ============================================================================
# Ribola CRAWLER CLASSES
# ============================================================================


logger = logging.getLogger(__name__)


class RibolaCrawler(BaseCrawler):
    """
    Crawler for Ribola store prices.

    This class handles downloading and parsing price data from Ribola's website.
    It fetches the HTML index page to find XML files for each store, downloads them,
    and parses the XML data to create a structured representation of stores and their products.
    """

    CHAIN = "ribola"
    BASE_URL = "https://ribola.hr"
    INDEX_URL = f"{BASE_URL}/ribola-cjenici/"

    # Known cities for address parsing
    CITIES = [
        "Kastel Sucurac",
        "Ploče",
        "Kaštel Gomilica",
        "Trogir",
        "Kaštel Lukšić",
        "Okrug Gornji",
        "Makarska",
        "Kaštel Stari",
        "Kaštel Novi",
        "Kastel Kambelovac",
        "Split",
        "Sinj",
        "Solin",
        "Orebić",
        "Nečujam",
        "Dubrovnik",
        "Podstrana",
        "Dugi Rat",
        "Ražanj",
        "Primošten",
        "Jelsa",
        "Stobrec",
        "Trilj",
        "Seget Donji",
        "Brela",
        "Šibenik",
        "Zadar",
    ]

    PRICE_MAP = {
        "price": ("MaloprodajnaCijena", False),
        "unit_price": ("CijenaZaJedinicuMjere", False),
        "special_price": ("MaloprodajnaCijenaAkcija", False),
        "best_price_30": ("NajnizaCijena", False),
        "anchor_price": ("SidrenaCijena", False),
    }

    FIELD_MAP = {
        "product": ("NazivProizvoda", True),
        "product_id": ("SifraProizvoda", True),
        "brand": ("MarkaProizvoda", False),
        "quantity": ("NetoKolicina", False),
        "unit": ("JedinicaMjere", False),
        "barcode": ("Barkod", False),
        "category": ("KategorijeProizvoda", False),
    }

    def parse_index(self, content: str) -> list[str]:
        """
        Parse the Ribola index page to extract XML file URLs.

        Args:
            content: HTML content of the index page

        Returns:
            List of XML file URLs found on the page
        """
        soup = BeautifulSoup(content, "html.parser")
        urls = []

        # Find all links ending with .xml
        for link_tag in soup.select('a[href$=".xml"]'):
            href = str(link_tag.get("href"))
            full_url = urljoin(self.INDEX_URL, href)
            urls.append(full_url)

        return list(set(urls))

    def parse_address_city(self, address_raw: str) -> tuple[str, str]:
        """
        Parse address and city from the combined string.

        Args:
            address_raw: Raw address string containing both street and city

        Returns:
            Tuple of (street_address, city)
        """
        address = address_raw.strip()

        # Check if it ends with any known city
        for city in self.CITIES:
            addr_norm = self.strip_diacritics(address.lower())
            city_norm = self.strip_diacritics(city.lower())

            if addr_norm.endswith(city_norm):
                # Strip city from the end to get street address
                street_address = address[: -len(city)].strip()
                return street_address, city

        # No known city found, treat entire string as address
        return address, ""

    def parse_store_info_from_xml(self, root: etree._Element) -> Store:
        """
        Parse store information from XML root element.

        Args:
            root: XML root element containing store data

        Returns:
            Store object with parsed store information
        """
        # Find the ProdajniObjekt element
        store_elem = root.find(".//ProdajniObjekt")
        if store_elem is None:
            raise ValueError("No ProdajniObjekt element found in XML")

        # Extract store information
        store_type_elem = store_elem.find("Oblik")
        store_type = (
            store_type_elem.text.lower()
            if store_type_elem is not None and store_type_elem.text
            else ""
        )

        store_id_elem = store_elem.find("Oznaka")
        store_id = (
            store_id_elem.text
            if store_id_elem is not None and store_id_elem.text
            else ""
        )

        address_elem = store_elem.find("Adresa")
        address_raw = (
            address_elem.text if address_elem is not None and address_elem.text else ""
        )

        street_address, city = self.parse_address_city(address_raw)

        store = Store(
            chain=self.CHAIN,
            store_type=store_type,
            store_id=store_id,
            name=f"{self.CHAIN.capitalize()} {city} {store_id}".strip(),
            street_address=street_address,
            zipcode="",
            city=city.title(),
            items=[],
        )

        logger.info(
            f"Parsed Ribola store: {store.name}, Type: {store.store_type}, "
            f"Address: {store.street_address}, City: {store.city}"
        )
        return store

    def parse_xml(self, xml_content: bytes) -> tuple[Store, list[Product]]:
        """
        Parse XML content into store info and list of products.

        Args:
            xml_content: XML content as bytes

        Returns:
            Tuple of (Store object, List of Product objects)
        """
        try:
            root = etree.fromstring(xml_content)

            # Parse store information
            store = self.parse_store_info_from_xml(root)

            # Parse products
            products = []
            for product_elem in root.xpath("//Proizvod"):
                try:
                    product = self.parse_xml_product(product_elem)
                    products.append(product)
                except Exception as e:
                    logger.warning(
                        f"Failed to parse product: {etree.tostring(product_elem)}: {e}",
                        exc_info=True,
                    )
                    continue

            logger.debug(f"Parsed {len(products)} products from XML")
            return store, products

        except Exception as e:
            logger.error(f"Failed to parse XML: {e}", exc_info=True)
            raise

    def get_store_data(self, xml_url: str) -> Store:
        """
        Fetch and parse both store info and products from a Ribola XML URL.

        Args:
            xml_url: URL to the XML file

        Returns:
            Store populated with Products
        """
        try:
            logger.debug(f"Fetching Ribola store data from: {xml_url}")

            xml_content = self.fetch_text(xml_url).encode("utf-8")
            store, products = self.parse_xml(xml_content)
            store.items = products
            return store
        except Exception as e:
            logger.error(
                f"Failed to get Ribola store data from {xml_url}: {e}",
                exc_info=True,
            )
            raise

    def get_index_urls_for_date(self, date: datetime.date) -> list[str]:
        """
        Fetch and parse the Ribola index page to get XML URLs for the specified date.

        Args:
            date: The date to search for in the price list.

        Returns:
            List of XML URLs containing data for the specified date.
        """
        index_url = f"{self.INDEX_URL}?date={date:%d.%m.%Y}"

        logger.debug(f"Fetching Ribola index page: {index_url}")

        content = self.fetch_text(index_url)
        if not content:
            logger.warning(f"No content found at Ribola index URL: {index_url}")
            return []

        xml_urls = list(set(self.parse_index(content)))

        if not xml_urls:
            logger.warning(f"No Ribola XML URLs found for date {date:%Y-%m-%d}")

        return xml_urls

    def get_all_products(self, date: datetime.date) -> list[Store]:
        """
        Main method to fetch and parse all Ribola store, product, and price info for a given date.

        Args:
            date: The date to search for in the price list.

        Returns:
            List of Store objects with their products.
        """
        xml_urls = self.get_index_urls_for_date(date)

        if not xml_urls:
            logger.warning(f"No Ribola XML URLs found for date {date.isoformat()}")
            return []

        stores = []
        for url in xml_urls:
            try:
                store = self.get_store_data(url)
            except Exception as e:
                logger.error(
                    f"Error processing Ribola store from {url}: {e}", exc_info=True
                )
                continue

            if not store.items:
                logger.warning(
                    f"No products found for Ribola store at {url}, skipping."
                )
                continue

            stores.append(store)

        return stores


if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG)
    crawler = RibolaCrawler()
    stores = crawler.crawl(datetime.date.today())
    print(stores[0])
    print(stores[0].items[0])


# ============================================================================
# Spar CRAWLER CLASSES
# ============================================================================


logger = logging.getLogger(__name__)


class SparCrawler(BaseCrawler):
    """
    Crawler for Spar/InterSpar store prices.

    This class handles downloading and parsing price data from Spar's website.
    It fetches the JSON index file, extracts CSV links, downloads and parses
    the CSVs, and returns a list of products.
    """

    CHAIN = "spar"
    BASE_URL = "https://www.spar.hr"
    ADDRESS_PATTERN = re.compile(
        r"^([a-zA-Z]+)_([a-zA-Z0-9_\.]+)_(\d{4,5})_([a-zA-Z_]+)_"
    )
    CITIES = [
        "varazdin",
        "valpovo",
        "sibenik",
        "zadar",
        "zagreb",
        "cakovec",
        "rijeka",
        "split",
        "kastav",
        "selce",
        "bibinje",
        "labin",
        "buje",
        "krizevci",
        "pozega",
        "jastrebarsko",
        "sesvetski_kraljevec",
        "krapinske_toplice",
        "novi_marof",
        "ivanic_grad",
        "vukovar",
        "marija_bistrica",
        "zapresic",
        "velika_gorica",
        "slavonski_brod",
        "osijek",
        "koprivnica",
        "bjelovar",
        "vinkovci",
        "dakovo",
        "orahovica",
        "pakrac",
        "suhopolje",
        "daruvar",
        "nasice",
        "pula",
        "opatija",
        "porec",
        "knin",
        "zlatar",
        "ivanec",
        "popovaca",
        "nin",
        "donja_stubica",
        "pregrada",
        "cepin",
        "ozalj",
        "dugo_selo",
        "gospic",
    ]
    PRICE_MAP = {
        "price": ("MPC", False),
        "unit_price": ("cijena za jedinicu mjere", False),
        "special_price": ("MPC za vrijeme posebnog oblika prodaje", False),
        "best_price_30": ("Najniža cijena u posljednjih 30 dana", False),
        "anchor_price": ("sidrena cijena na 2.5.2025.", False),
    }

    FIELD_MAP = {
        "barcode": ("barkod", False),
        "product": ("naziv", True),
        "product_id": ("šifra", True),
        "brand": ("marka", False),
        "quantity": ("neto količina", False),
        "unit": ("jedinica mjere", False),
        "category": ("kategorija proizvoda", False),
        "anchor_price_date": ("datum sidrene cijene", False),
    }

    # Required to detect text encoding
    CSV_PREFIX = "naziv;šifra;marka;neto količina;jedinica mjere;"

    def fetch_price_list_index(self, date: datetime.date) -> dict[str, str]:
        """
        Fetch the JSON index file with list of CSV files.

        Args:
            date: The date for which to fetch the price list index

        Returns:
            A dictionary with filename → URL mappings for CSV files.

        Raises:
            httpx.RequestError: If the request fails
        """
        url = f"{self.BASE_URL}/datoteke_cjenici/Cjenik{date:%Y%m%d}.json"
        content = self.fetch_text(url)

        json_data = loads(content)
        files = json_data.get("files")
        if not files:
            logger.error("Price list index doesn't contain any files")
            return {}

        return {info.get("name", ""): info.get("URL", "") for info in files}

    def parse_store_from_filename(self, filename: str) -> Optional[Store]:
        """
        Extract store information from CSV filename using regex.

        Supported filename pattern:
            `hipermarket_zadar_bleiburskih_zrtava_18_8701_interspar_zadar_0017_20250518_0330.csv`

        Args:
            filename: Name of the CSV file with store information

        Returns:
            Store object with parsed store information, or None if parsing fails
        """
        logger.debug(f"Parsing store information from filename: {filename}")

        match = self.ADDRESS_PATTERN.match(filename)

        if not match:
            logger.warning(f"Failed to match filename pattern: {filename}")
            return None

        store_type, city_and_address, store_id, store_name = match.groups()

        for city in self.CITIES:
            if city_and_address.lower().startswith(city):
                store_city = city
                store_address = city_and_address[len(city) + 1 :]
                break
        else:
            # Assume city is the first word
            store_city, store_address = city_and_address.split("_", 1)

        store = Store(
            chain="spar",
            store_id=store_id,
            name=store_name.replace("_", " ").title(),
            store_type=store_type.lower(),
            city=store_city.replace("_", " ").title(),
            street_address=store_address.replace("_", " ").title(),
            items=[],
        )

        logger.debug(
            f"Parsed store: {store.name} ({store.store_id}), {store.store_type}, {store.city}, {store.street_address}"
        )
        return store

    def parse_csv_row(self, row: dict) -> Product:
        fixed_row = {k.replace("(EUR)", "").strip(): v for k, v in row.items()}
        return super().parse_csv_row(fixed_row)

    def get_all_products(self, date: datetime.date) -> list[Store]:
        """
        Main method to fetch and parse all products from Spar's price lists.

        Args:
            date: The date for which to fetch the price list

        Returns:
            Tuple with the date and the list of Store objects,
            each containing its products.

        Raises:
            ValueError: If the price list index cannot be fetched or parsed
        """
        # Fetch the price list index
        csv_files = self.fetch_price_list_index(date)

        logger.info(f"Found {len(csv_files)} CSV files in the price list index")

        stores = []

        for filename, url in csv_files.items():
            store = self.parse_store_from_filename(filename)
            if not store:
                logger.warning(f"Skipping CSV from {url} due to store parsing failure")
                continue

            csv_content = self.fetch_text(
                url, ["iso-8859-2", "windows-1250"], self.CSV_PREFIX
            )
            if not csv_content:
                logger.warning(f"Skipping CSV from {url} due to download failure")
                continue

            try:
                products = self.parse_csv(csv_content, ";")
                store.items = products
                stores.append(store)
            except Exception as e:
                logger.error(f"Error processing CSV from {url}: {e}", exc_info=True)
                continue

        return stores


if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG)
    crawler = SparCrawler()
    stores = crawler.get_all_products(datetime.date.today())
    print(stores[0])
    print(stores[0].items[0])



# ============================================================================
# Studenac CRAWLER CLASSES
# ============================================================================


logger = logging.getLogger(__name__)


class StudenacCrawler(BaseCrawler):
    """
    Crawler for Studenac store prices.

    This class handles downloading and parsing price data from Studenac's website.
    It fetches the ZIP file containing XML files for each store, extracts them,
    and parses the XML data to create a structured representation of stores and their products.
    """

    CHAIN = "studenac"
    BASE_URL = "https://www.studenac.hr"
    TIMEOUT = 120.0  # Longer timeout for ZIP download

    PRICE_MAP = {
        "price": ("MaloprodajnaCijena", False),
        "unit_price": ("CijenaPoJedinici", False),
        "special_price": ("MaloprodajnaCijenaAkcija", False),
        "best_price_30": ("NajnizaCijena", False),
        "anchor_price": ("SidrenaCijena", False),
    }

    FIELD_MAP = {
        "product": ("NazivProizvoda", False),
        "product_id": ("SifraProizvoda", True),
        "brand": ("MarkaProizvoda", False),
        "quantity": ("NetoKolicina", False),
        "unit": ("JedinicaMjere", False),
        "barcode": ("Barkod", False),
        "category": ("KategorijeProizvoda", False),
    }

    def parse_address(self, address: str) -> Tuple[str, str]:
        """
        Parse the address string into street address and city components.

        Args:
            address: Address string in format "<street> <number> <CITY>"

        Returns:
            Tuple of (street_address, city)
        """
        logger.debug(f"Parsing address: {address}")

        try:
            # The regex matches the last set of uppercase words (city)
            # and everything before it (street address)
            pattern = r"^(.*?)([A-ZČĆĐŠŽ][A-ZČĆĐŠŽ\s]+)$"
            match = re.match(pattern, address)

            if match:
                street_address, city = match.groups()
                return (
                    street_address.strip().title(),
                    city.strip().title(),
                )

            logger.warning(f"Failed to parse address: {address}")
            return address.strip().title(), ""
        except Exception as e:
            logger.warning(f"Error parsing address {address}: {e}", exc_info=True)
            return address.strip().title(), ""

    def parse_xml(self, xml_content: bytes) -> Optional[Store]:
        """
        Parse XML content into a unified Store object.

        Args:
            xml_content: XML content as bytes

        Returns:
            Store object with parsed store and product information,
            or None if parsing fails
        """
        try:
            root = etree.fromstring(xml_content)

            # Extract store information
            store_type = root.xpath("//ProdajniObjekt/Oblik/text()")[0].lower()
            store_id = root.xpath("//ProdajniObjekt/Oznaka/text()")[0]
            store_code = root.xpath("//ProdajniObjekt/Oznaka/text()")[0]
            address = root.xpath("//ProdajniObjekt/Adresa/text()")[0]

            street_address, city = self.parse_address(address)

            store = Store(
                chain=self.CHAIN,
                name=f"Studenac {store_code}",
                store_type=store_type.lower(),
                store_id=store_id,
                city=city,
                street_address=street_address,
                items=[],
            )

            logger.debug(
                f"Parsed store: {store.name} ({store_id}), {store.store_type}, {store.city}, {store.street_address}"
            )

            # Extract product information
            products = []
            for product_elem in root.xpath("//ProdajniObjekt/Proizvodi/Proizvod"):
                try:
                    product = self.parse_xml_product(product_elem)
                    products.append(product)
                except Exception as e:
                    logger.warning(
                        f"Failed to parse product: {etree.tostring(product_elem)}: {e}",
                        exc_info=True,
                    )
                    continue

            store.items = products
            logger.debug(f"Parsed {len(products)} products for store {store.name}")
            return store

        except Exception as e:
            logger.error(f"Failed to parse XML: {e}", exc_info=True)
            return None

    def get_all_products(self, date: datetime.date) -> list[Store]:
        """
        Main method to fetch and parse all products from Studenac's price lists.

        Args:
            date: The date for which to fetch the price list

        Returns:
            Tuple with the date and the list of Store objects,
            each containing its products.

        Raises:
            ValueError: If the price list cannot be fetched or parsed
        """
        stores = []
        zip_url = f"{self.BASE_URL}/cjenici/PROIZVODI-{date:%Y-%m-%d}.zip"

        for filename, content in self.get_zip_contents(zip_url, ".xml"):
            logger.debug(f"Processing file: {filename}")
            store = self.parse_xml(content)
            if store:
                stores.append(store)

        return stores

    def get_zip_contents(
        self, url: str, suffix: str
    ) -> Generator[tuple[str, bytes], None, None]:
        with TemporaryDirectory() as temp_dir:  # type: ignore
            temp_path = Path(temp_dir)
            temp_zip = temp_path / "archive.zip"
            with open(temp_zip, "wb") as fp:
                self.fetch_binary(url, fp)

            subprocess.run(["unzip", "-x", temp_zip], cwd=temp_dir)

            for file in temp_path.iterdir():
                if file.suffix != suffix:
                    continue

                xml_content = open(file, "rb").read()
                yield file.name, xml_content


if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG)
    crawler = StudenacCrawler()
    stores = crawler.crawl(datetime.date.today())
    print(stores[0])
    print(stores[0].items[0])


# ============================================================================
#  Tommy CRAWLER CLASSES
# ============================================================================


logger = logging.getLogger(__name__)


class TommyCrawler(BaseCrawler):
    """
    Crawler for Tommy store prices.

    This class handles downloading and parsing price data from Tommy's API.
    It retrieves JSON data about available store price tables and processes
    the corresponding CSV files for product information.
    """

    CHAIN = "tommy"
    BASE_URL = "https://spiza.tommy.hr/api/v2"

    def fetch_stores_list(self, date: datetime.date) -> dict[str, str]:
        """
        Fetch the list of store price tables for a specific date.

        Args:
            date: The date for which to fetch the price tables

        Returns:
            List of dictionaries containing store price table information

        Raises:
            httpx.RequestError: If the API request fails
            ValueError: If the response cannot be parsed
        """
        url = (
            f"{self.BASE_URL}/shop/store-prices-tables"
            f"?date={date:%Y-%m-%d}&page=1&itemsPerPage=200&channelCode=general"
        )
        content = self.fetch_text(url)
        data = loads(content)
        store_list = data.get("hydra:member", [])

        stores = {}
        for store in store_list:
            csv_id = store.get("@id")
            filename = store.get("fileName", "Unknown")
            if not csv_id or not filename:
                logger.warning(
                    f"Skipping store with missing CSV ID or filename: {store}"
                )
                continue
            if csv_id.startswith("/api/v2"):
                csv_id = csv_id[len("/api/v2") :]

            stores[filename] = self.BASE_URL + csv_id

        return stores

    def parse_date_string(self, date_str: str) -> Optional[datetime.date]:
        """
        Parse date string from CSV (format DD.MM.YYYY. HH:MM:SS).

        Args:
            date_str: The date string to parse (e.g., "16.5.2025. 0:00:00")

        Returns:
            datetime.date object or None if parsing fails
        """
        if not date_str or date_str.strip() == "":
            return None

        try:
            # Use regex to extract day, month, and year
            # The pattern handles both single and double-digit day/month
            match = re.match(r"(\d{1,2})\.(\d{1,2})\.(\d{4})\.", date_str)

            if match:
                day, month, year = map(int, match.groups())
                return datetime.date(year, month, day)
            else:
                logger.warning(f"Date string format not recognized: {date_str}")
                return None

        except (ValueError, IndexError) as e:
            logger.warning(f"Failed to parse date string '{date_str}': {e}")
            return None

    def parse_csv(self, csv_content: str) -> List[Product]:
        """
        Parse CSV content and extract product information.

        Args:
            csv_content: Content of the CSV file

        Returns:
            List of Product objects

        CSV format:
            BARKOD_ARTIKLA,SIFRA_ARTIKLA,NAZIV_ARTIKLA,BRAND,ROBNA_STRUKTURA,
            JEDINICA_MJERE,NETO_KOLICINA,MPC,MPC_POSEBNA_PRODAJA,CIJENA_PO_JM,
            MPC_NAJNIZA_30,MPC_020525,DATUM_ULASKA_NOVOG_ARTIKLA,PRVA_CIJENA_NOVOG_ARTIKLA
        """
        logger.debug("Parsing CSV content")

        products = []
        success_count = 0
        error_count = 0

        try:
            # Read CSV content using StringIO and DictReader
            csv_file = io.StringIO(csv_content)
            reader = csv.DictReader(csv_file)  # type: ignore

            if not reader.fieldnames:
                logger.warning("CSV file has no header row")
                return products

            logger.debug(f"CSV header: {reader.fieldnames}")

            # Define expected field names
            field_map = {
                "barcode": "BARKOD_ARTIKLA",
                "product_id": "SIFRA_ARTIKLA",
                "product_name": "NAZIV_ARTIKLA",
                "brand": "BRAND",
                "category": "ROBNA_STRUKTURA",
                "unit": "JEDINICA_MJERE",
                "quantity": "NETO_KOLICINA",
                "price": "MPC",
                "special_price": "MPC_POSEBNA_PRODAJA",
                "unit_price": "CIJENA_PO_JM",
                "lowest_price_30days": "MPC_NAJNIZA_30",
                "anchor_price": "MPC_020525",
                "date_added": "DATUM_ULASKA_NOVOG_ARTIKLA",
                "initial_price": "PRVA_CIJENA_NOVOG_ARTIKLA",
            }

            row_count = 0
            for row in reader:
                row_count += 1

                try:
                    # Extract mandatory fields from the row
                    barcode = row.get(field_map["barcode"], "").strip()
                    product_id = row.get(field_map["product_id"], "").strip()
                    product_name = row.get(field_map["product_name"], "").strip()
                    brand = row.get(field_map["brand"], "").strip()
                    category = row.get(field_map["category"], "").strip()
                    unit = row.get(field_map["unit"], "").strip()
                    quantity = row.get(field_map["quantity"], "").strip()

                    # Parse price fields with proper error handling
                    try:
                        price = parse_price(row.get(field_map["price"], "0"))
                    except Exception as e:
                        logger.warning(f"Failed to parse price in row {row_count}: {e}")
                        price = Decimal("0.00")

                    try:
                        unit_price = parse_price(row.get(field_map["unit_price"], "0"))
                    except Exception as e:
                        logger.warning(
                            f"Failed to parse unit_price in row {row_count}: {e}"
                        )
                        unit_price = Decimal("0.00")

                    # Parse optional price fields
                    special_price = None
                    lowest_price_30days = None
                    anchor_price = None
                    initial_price = None
                    date_added = None

                    special_price_str = row.get(field_map["special_price"], "")
                    if special_price_str.strip():
                        try:
                            special_price = parse_price(special_price_str)
                        except Exception:
                            pass

                    lowest_price_30days_str = row.get(
                        field_map["lowest_price_30days"], ""
                    )
                    if lowest_price_30days_str.strip():
                        try:
                            lowest_price_30days = parse_price(lowest_price_30days_str)
                        except Exception:
                            pass

                    anchor_price_str = row.get(field_map["anchor_price"], "")
                    if anchor_price_str.strip():
                        try:
                            anchor_price = parse_price(anchor_price_str)
                        except Exception:
                            pass

                    date_added_str = row.get(field_map["date_added"], "")
                    if date_added_str.strip():
                        date_added = self.parse_date_string(date_added_str)

                    initial_price_str = row.get(field_map["initial_price"], "")
                    if initial_price_str.strip():
                        try:
                            initial_price = parse_price(initial_price_str)
                        except Exception:
                            pass

                    # Create product if we have the minimum required fields
                    if product_name and (price or unit_price):
                        # If one price is missing but the other exists, use the existing one for both
                        if price and not unit_price:
                            unit_price = price
                        elif unit_price and not price:
                            price = unit_price

                        product = Product(
                            product=product_name,
                            product_id=product_id,
                            barcode=barcode,
                            brand=brand,
                            category=category,
                            unit=unit,
                            quantity=quantity,
                            price=price,
                            special_price=special_price,
                            unit_price=unit_price,
                            best_price_30=lowest_price_30days,  # Map lowest_price_30days to best_price_30
                            anchor_price=anchor_price,
                            date_added=date_added,
                            initial_price=initial_price,
                        )
                        products.append(product)
                        success_count += 1
                    else:
                        logger.warning(
                            f"Skipping product in row {row_count} with missing required fields: {row}"
                        )
                        error_count += 1

                except Exception as e:
                    logger.error(f"Error parsing product row {row_count}: {e}")
                    logger.debug(f"Problematic row: {row}")
                    error_count += 1

            logger.info(
                f"Parsed {len(products)} products from CSV (total rows: {row_count}, errors: {error_count})"
            )
            return products

        except Exception as e:
            logger.error(f"Error parsing CSV: {e}")
            return []

    def parse_store_from_filename(
        self, filename: str
    ) -> Tuple[str, str, str, str, str]:
        """
        Parse store information from the filename.

        Args:
            filename: The filename from the API

        Returns:
            Tuple of (store_type, address, zipcode, city)

        Example:
            "SUPERMARKET, ANTE STARČEVIĆA 6, 20260 KORČULA, 10180, 2, 20250516 0530"
            Will return:
            ("supermarket", "10180", "Ante Starčevića 6", "20260", "Korčula")
        """
        try:
            # Split by commas
            parts = filename.split(",")

            if len(parts) < 3:
                logger.warning(f"Filename doesn't have enough parts: {filename}")
                raise ValueError(f"Unparseable filename: {filename}")

            # Extract store type (first part)
            store_type = parts[0].strip().lower()

            # Extract address (second part)
            address = to_camel_case(parts[1].strip())

            # Extract zipcode and city (third part)
            location_part = parts[2].strip()

            # Use regex to extract zipcode and city
            # Pattern looks for 5 digits followed by any text
            match = re.match(r"(\d{5})\s+(.+)", location_part)

            if match:
                zipcode = match.group(1)
                city = to_camel_case(match.group(2))
            else:
                logger.warning(
                    f"Could not extract zipcode and city from: {location_part}"
                )
                zipcode = ""
                # Try to extract just the city if no zipcode pattern found
                city = to_camel_case(location_part)

            store_id = parts[3].strip()

            logger.debug(
                f"Parsed store info: type={store_type}, address={address}, zipcode={zipcode}, city={city}"
            )

            return (store_type, store_id, address, zipcode, city)

        except Exception as e:
            logger.error(f"Error parsing store from filename {filename}: {e}")
            raise

    def get_all_products(self, date: datetime.date) -> list[Store]:
        """
        Main method to fetch and parse all products from Tommy's price lists.

        Args:
            date: The date for which to fetch the price list

        Returns:
            Tuple with the date and the list of Store objects,
            each containing its products.

        Raises:
            ValueError: If the price list cannot be fetched or parsed
        """

        store_map = self.fetch_stores_list(date)
        if not store_map:
            logger.warning(f"No stores found for date {date}")
            return []

        stores = []
        for filename, url in store_map.items():
            # Extract store information
            store_type, store_id, address, zipcode, city = (
                self.parse_store_from_filename(filename)
            )

            store = Store(
                chain="tommy",
                name=f"Tommy {store_type.title()} {address}",
                store_type=store_type,
                store_id=store_id,
                city=city,
                street_address=address,
                zipcode=zipcode,
                items=[],
            )

            csv_content = self.fetch_text(url)
            products = self.parse_csv(csv_content)

            store.items = products
            stores.append(store)

        return stores


if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG)
    crawler = TommyCrawler()
    current_date = datetime.date.today() - datetime.timedelta(days=1)
    stores = crawler.get_all_products(current_date)
    print(stores[0])
    print(stores[0].items[0])


# ============================================================================
#  Trgocentar CRAWLER CLASSES
# ============================================================================


logger = logging.getLogger(__name__)


class TrgocentarCrawler(BaseCrawler):
    """
    Crawler for Trgocentar store prices.

    This class handles downloading and parsing price data from Trgocentar's website.
    It fetches the HTML index page to find XML files for each store, downloads them,
    and parses the XML data to create a structured representation of stores and their products.
    """

    CHAIN = "trgocentar"
    BASE_URL = "https://trgocentar.com"
    INDEX_URL = "https://trgocentar.com/Trgovine-cjenik/"

    # Regex to parse store information from XML filename
    # Format: <store_type>_<address_parts>_P<store_id>_<serial>_<DDMMYYYY><time>.xml
    # Example: SUPERMARKET_VL_NAZORA_58_SV_IVAN_ZELINA_P120_009_230520250745.xml
    FILENAME_PATTERN = re.compile(
        r"^(?P<store_type>[^_]+)_"
        r"(?P<address_city>.+?)_"
        r"P(?P<store_id>\d+)_"
        r"(?P<serial>\d+)_"
        r"(?P<date>\d{8})"
        r"(?P<time>\d+)\.xml$"
    )

    # Known cities to detect and separate from address
    CITIES = [
        "HUM NA SUTLI",
        "ZLATAR",
        "SV IVAN ZELINA",
        "SV KRIZ ZACRETJE",
        "ZABOK",
        "ZAPRESIC",
    ]

    PRICE_MAP = {
        "price": ("mpc", False),
        "unit_price": ("c_jmj", False),
        "special_price": ("mpc_pop", False),
        "best_price_30": ("c_najniza_30", False),
        "anchor_price": ("c_020525", False),
    }

    FIELD_MAP = {
        "product": ("naziv_art", True),
        "product_id": ("sif_art", True),
        "brand": ("marka", False),
        "quantity": ("net_kol", False),
        "unit": ("jmj", False),
        "barcode": ("ean_kod", False),
        "category": ("naz_kat", False),
    }

    def parse_index(self, content: str) -> list[str]:
        """
        Parse the Trgocentar index page to extract XML file URLs.

        Args:
            content: HTML content of the index page

        Returns:
            List of XML file URLs found on the page
        """
        soup = BeautifulSoup(content, "html.parser")
        urls = []

        # Find all links ending with .xml
        for link_tag in soup.select('a[href$=".xml"]'):
            href = str(link_tag.get("href"))
            full_url = urljoin(self.INDEX_URL, href)
            urls.append(full_url)

        return list(set(urls))

    def parse_address_city(self, address_city_raw: str) -> tuple[str, str]:
        """
        Parse address and city from the combined string.

        Args:
            address_city_raw: Raw address+city string with underscores

        Returns:
            Tuple of (street_address, city)
        """
        # Convert underscores to spaces
        address_city = address_city_raw.replace("_", " ")

        # Check if it ends with any known city
        for city in self.CITIES:
            if address_city.endswith(city):
                # Strip city from the end to get address
                street_address = address_city[: -len(city)].strip()
                return street_address.title(), city.title()

        # No known city found, treat entire string as address
        return address_city.title(), ""

    def parse_store_info(self, xml_url: str) -> Store:
        """
        Parse store information from an XML file URL.

        Args:
            xml_url: URL to the XML file containing store/product data

        Returns:
            Store object with parsed store information
        """
        logger.debug(f"Parsing store information from Trgocentar URL: {xml_url}")

        filename = os.path.basename(xml_url)
        match = self.FILENAME_PATTERN.match(filename)

        if not match:
            raise ValueError(f"Invalid XML filename format for Trgocentar: {filename}")

        data = match.groupdict()

        store_type = data["store_type"].lower()
        store_id = f"P{data['store_id']}"
        street_address, city = self.parse_address_city(data["address_city"])

        store = Store(
            chain=self.CHAIN,
            store_type=store_type,
            store_id=store_id,
            name=f"{self.CHAIN.capitalize()} {city} {store_id}".strip(),
            street_address=street_address,
            zipcode="",
            city=city,
            items=[],
        )

        logger.info(
            f"Parsed Trgocentar store: {store.name}, Type: {store.store_type}, "
            f"Address: {store.street_address}, City: {store.city}"
        )
        return store

    def parse_xml(self, xml_content: bytes) -> list[Product]:
        """
        Parse XML content into a list of products.

        Args:
            xml_content: XML content as bytes

        Returns:
            List of Product objects parsed from the XML
        """
        try:
            root = etree.fromstring(xml_content)
            products = []

            for product_elem in root.xpath("//cjenik"):
                try:
                    product = self.parse_xml_product(product_elem)
                    products.append(product)
                except Exception as e:
                    logger.warning(
                        f"Failed to parse product: {etree.tostring(product_elem)}: {e}",
                        exc_info=True,
                    )
                    continue

            logger.debug(f"Parsed {len(products)} products from XML")
            return products

        except Exception as e:
            logger.error(f"Failed to parse XML: {e}", exc_info=True)
            return []

    def get_store_data(self, xml_url: str) -> Store:
        """
        Fetch and parse both store info and products from a Trgocentar XML URL.

        Args:
            xml_url: URL to the XML file

        Returns:
            Store populated with with Products
        """
        try:
            store = self.parse_store_info(xml_url)

            xml_content = self.fetch_text(xml_url).encode("utf-8")
            products = self.parse_xml(xml_content)
            store.items = products
            return store
        except Exception as e:
            logger.error(
                f"Failed to get Trgocentar store data from {xml_url}: {e}",
                exc_info=True,
            )
            raise

    def get_index_urls_for_date(self, date: datetime.date) -> list[str]:
        """
        Fetch and parse the Trgocentar index page to get XML URLs for the specified date.

        Args:
            date: The date to search for in the XML filenames (DDMMYYYY format).

        Returns:
            List of XML URLs containing data for the specified date.
        """
        content = self.fetch_text(self.INDEX_URL)

        if not content:
            logger.warning(
                f"No content found at Trgocentar index URL: {self.INDEX_URL}"
            )
            return []

        all_urls = self.parse_index(content)

        # Date format in Trgocentar filenames is DDMMYYYY
        date_str = date.strftime("%d%m%Y")

        matching_urls = []
        for url in all_urls:
            filename = os.path.basename(url)
            # Check if the DDMMYYYY date string is in the filename with underscore prefix
            if f"_{date_str}" in filename:
                matching_urls.append(url)

        if not matching_urls:
            logger.warning(f"No Trgocentar URLs found matching date {date:%Y-%m-%d}")

        return matching_urls

    def get_all_products(self, date: datetime.date) -> list[Store]:
        """
        Main method to fetch and parse all Trgocentar store, product, and price info for a given date.

        Args:
            date: The date to search for in the price list.

        Returns:
            List of Store objects with their products.
        """
        xml_urls = self.get_index_urls_for_date(date)

        if not xml_urls:
            logger.warning(f"No Trgocentar XML URLs found for date {date.isoformat()}")
            return []

        stores = []
        for url in xml_urls:
            try:
                store = self.get_store_data(url)
            except Exception as e:
                logger.error(
                    f"Error processing Trgocentar store from {url}: {e}", exc_info=True
                )
                continue

            if not store.items:
                logger.warning(
                    f"No products found for Trgocentar store at {url}, skipping."
                )
                continue

            stores.append(store)

        return stores


if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG)
    crawler = TrgocentarCrawler()
    stores = crawler.crawl(datetime.date.today())
    print(stores[0])
    print(stores[0].items[0])

# ============================================================================
#  Vrutak CRAWLER CLASSES
# ============================================================================


logger = logging.getLogger(__name__)


class VrutakCrawler(BaseCrawler):
    """
    Crawler for Vrutak store prices.

    This class handles downloading and parsing price data from Vrutak's website.
    It fetches the HTML index page to find XML files for each store, downloads them,
    and parses the XML data to create a structured representation of stores and their products.
    """

    CHAIN = "vrutak"
    BASE_URL = "https://www.vrutak.hr"
    INDEX_URL = "https://www.vrutak.hr/cjenik-svih-artikala"

    # Known store types
    STORE_TYPES = ["hipermarket", "supermarket"]

    PRICE_MAP = {
        "price": ("mpcijena", True),
        "unit_price": ("mpcijenamjera", False),
        "special_price": ("", False),  # No equivalent in Vrutak XML
        "best_price_30": ("", False),  # No equivalent in Vrutak XML
        "anchor_price": ("", False),  # No equivalent in Vrutak XML
    }

    FIELD_MAP = {
        "product": ("naziv", True),
        "product_id": ("sifra", True),
        "brand": ("marka", False),
        "quantity": ("nettokolicina", False),
        "unit": ("mjera", False),
        "barcode": ("barkod", False),
        "category": ("kategorija", False),
    }

    def parse_index(self, content: str) -> dict[datetime.date, list[str]]:
        """
        Parse the Vrutak index page to extract XML file URLs grouped by date.

        Args:
            content: HTML content of the index page

        Returns:
            Dictionary mapping dates to lists of XML file URLs
        """
        soup = BeautifulSoup(content, "html.parser")
        urls_by_date = {}

        # Find all rows in tbody
        for row in soup.select("tbody tr"):
            cells = row.select("td")
            if len(cells) < 3:
                continue

            # Second cell contains the date
            date_cell = cells[1]
            date_text = date_cell.get_text(strip=True)

            try:
                # Parse date in DD.MM.YYYY format
                date_obj = datetime.datetime.strptime(date_text, "%d.%m.%Y.").date()
            except ValueError:
                # Non-data row
                continue

            # Extract XML URLs from remaining cells
            xml_urls = []
            for cell in cells[2:]:  # Skip index and date cells
                link = cell.select_one('a[href$=".xml"]')
                if link:
                    href = str(link.get("href"))
                    full_url = urljoin(self.BASE_URL, href)
                    xml_urls.append(full_url)

            if xml_urls:
                urls_by_date[date_obj] = xml_urls

        return urls_by_date

    def parse_store_info(self, xml_url: str) -> Store:
        """
        Parse store information from an XML file URL.

        Args:
            xml_url: URL to the XML file containing store/product data

        Returns:
            Store object with parsed store information
        """
        logger.debug(f"Parsing store information from Vrutak URL: {xml_url}")

        filename = os.path.basename(xml_url)
        # Remove .xml extension and split by dashes
        parts = filename[:-4].split("-")

        if len(parts) < 4:
            raise ValueError(f"Invalid XML filename format for Vrutak: {filename}")

        # Expected format: vrutak-type-address-store_id-serial-datetime
        store_type = parts[1]  # hipermarket or supermarket
        street_address = parts[2].title()
        store_id = parts[3]

        store = Store(
            chain=self.CHAIN,
            store_type=store_type,
            store_id=store_id,
            name=f"{self.CHAIN.capitalize()} {store_type} {store_id}",
            street_address=street_address,
            zipcode="10000",
            city="Zagreb",
            items=[],
        )

        logger.info(
            f"Parsed Vrutak store: {store.name}, Type: {store.store_type}, "
            f"Address: {store.street_address}, City: {store.city}"
        )
        return store

    def parse_xml(self, xml_content: bytes) -> list[Product]:
        """
        Parse XML content into a list of products.

        Args:
            xml_content: XML content as bytes

        Returns:
            List of Product objects parsed from the XML
        """
        try:
            root = etree.fromstring(xml_content)
            products = []

            for product_elem in root.xpath("//item"):
                try:
                    product = self.parse_xml_product(product_elem)
                    products.append(product)
                except Exception as e:
                    logger.warning(
                        f"Failed to parse product: {etree.tostring(product_elem)}: {e}",
                        exc_info=True,
                    )
                    continue

            logger.debug(f"Parsed {len(products)} products from XML")
            return products

        except Exception as e:
            logger.error(f"Failed to parse XML: {e}", exc_info=True)
            return []

    def get_store_data(self, xml_url: str) -> Store:
        """
        Fetch and parse both store info and products from a Vrutak XML URL.

        Args:
            xml_url: URL to the XML file

        Returns:
            Store populated with Products
        """
        try:
            store = self.parse_store_info(xml_url)

            xml_content = self.fetch_text(xml_url).encode("utf-8")
            products = self.parse_xml(xml_content)
            store.items = products
            return store
        except Exception as e:
            logger.error(
                f"Failed to get Vrutak store data from {xml_url}: {e}",
                exc_info=True,
            )
            raise

    def get_index_urls_for_date(self, date: datetime.date) -> list[str]:
        """
        Fetch and parse the Vrutak index page to get XML URLs for the specified date.

        Args:
            date: The date to search for in the XML filenames.

        Returns:
            List of XML URLs containing data for the specified date.
        """
        content = self.fetch_text(self.INDEX_URL)

        if not content:
            logger.warning(f"No content found at Vrutak index URL: {self.INDEX_URL}")
            return []

        urls_by_date = self.parse_index(content)
        matching_urls = urls_by_date.get(date, [])

        if not matching_urls:
            logger.warning(f"No Vrutak URLs found matching date {date:%Y-%m-%d}")

        return matching_urls

    def get_all_products(self, date: datetime.date) -> list[Store]:
        """
        Main method to fetch and parse all Vrutak store, product, and price info for a given date.

        Args:
            date: The date to search for in the price list.

        Returns:
            List of Store objects with their products.
        """
        xml_urls = self.get_index_urls_for_date(date)

        if not xml_urls:
            logger.warning(f"No Vrutak XML URLs found for date {date.isoformat()}")
            return []

        stores = []
        for url in xml_urls:
            try:
                store = self.get_store_data(url)
            except Exception as e:
                logger.error(
                    f"Error processing Vrutak store from {url}: {e}", exc_info=True
                )
                continue

            if not store.items:
                logger.warning(
                    f"No products found for Vrutak store at {url}, skipping."
                )
                continue

            stores.append(store)

        return stores


if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG)
    crawler = VrutakCrawler()
    stores = crawler.crawl(datetime.date.today())
    print(stores[0])
    print(stores[0].items[0])


# ============================================================================
#  Zabac CRAWLER CLASSES
# ============================================================================

rom .base import BaseCrawler

logger = logging.getLogger(__name__)


class ZabacCrawler(BaseCrawler):
    """Crawler for Žabac store prices."""

    CHAIN = "zabac"
    BASE_URL = "https://zabacfoodoutlet.hr/cjenik/"

    # Regex to parse store information from the filename
    # Format: Cjenik-Zabac-Food-Outlet-PJ-<store_id>-<address>.csv
    # Example: Cjenik-Zabac-Food-Outlet-PJ-11-Savska-Cesta-206.csv
    STORE_FILENAME_PATTERN = re.compile(r".*PJ-(?P<store_id>\d+)-(?P<address>.+)\.csv$")

    # Mapping for price fields from CSV columns
    PRICE_MAP = {
        # field: (column_name, is_required)
        "price": ("MPC", False),
        "unit_price": ("MPC", False),  # Use same as price
        "special_price": ("", False),  # Not available in Zabac CSV
        "anchor_price": ("", False),  # Not available in Zabac CSV
    }

    # Mapping for other product fields from CSV columns
    FIELD_MAP = {
        "product_id": ("Artikl Šifra", True),
        "barcode": ("Barcode", False),
        "product": ("Naziv artikla / usluge", True),
        "brand": ("", False),  # Not available in Zabac CSV
        "quantity": ("", False),  # Not available in Zabac CSV
        "unit": ("", False),  # Not available in Zabac CSV
        "category": ("", False),  # Not available in Zabac CSV
    }

    def parse_index(self, content: str) -> list[str]:
        """
        Parse the Žabac index page to extract CSV links.

        Args:
            content: HTML content of the index page

        Returns:
            List of absolute CSV URLs on the page
        """
        soup = BeautifulSoup(content, "html.parser")
        urls = []

        for link_tag in soup.select('a[href$=".csv"]'):
            href = str(link_tag.get("href"))
            urls.append(href)

        return list(set(urls))  # Return unique URLs

    def parse_store_info(self, url: str) -> Store:
        """
        Extracts store information from a CSV download URL.

        Example URL:
        https://zabacfoodoutlet.hr/wp-content/uploads/2025/05/Cjenik-Zabac-Food-Outlet-PJ-11-Savska-Cesta-206.csv

        Args:
            url: CSV download URL with store information in the filename

        Returns:
            Store object with parsed store information
        """
        logger.debug(f"Parsing store information from Zabac URL: {url}")

        filename = unquote(os.path.basename(url))

        match = self.STORE_FILENAME_PATTERN.match(filename)
        if not match:
            raise ValueError(f"Invalid CSV filename format for Zabac: {filename}")

        data = match.groupdict()

        store_id = data["store_id"]
        # Address: "Savska-Cesta-206" -> "Savska Cesta 206"
        address_raw = data["address"]
        street_address = address_raw.replace("-", " ")

        store = Store(
            chain=self.CHAIN,
            store_type="",  # Store type is not available in the filename
            store_id=f"PJ-{store_id}",
            name=f"Žabac PJ-{store_id}",  # e.g. "Žabac PJ-11"
            street_address=street_address,
            zipcode="",  # Zipcode is not available in the filename
            city="",  # City is not available in the filename
            items=[],
        )

        logger.info(
            f"Parsed Žabac store: {store.name}, Address: {store.street_address}"
        )
        return store

    def get_store_prices(self, csv_url: str) -> list[Product]:
        """
        Fetch and parse store prices from a Žabac CSV URL.
        The CSV is semicolon-separated and windows-1250 encoded.

        Args:
            csv_url: URL to the CSV file containing prices

        Returns:
            List of Product objects
        """
        try:
            content = self.fetch_text(csv_url, encodings=["windows-1250"])
            return self.parse_csv(content, delimiter=";")
        except Exception as e:
            logger.error(
                f"Failed to get Žabac store prices from {csv_url}: {e}",
                exc_info=True,
            )
            return []

    def get_index(self, date: datetime.date) -> list[str]:
        """
        Fetch and parse the Žabac index page to get CSV URLs.

        Note: Žabac only shows current CSV files, so the date parameter is ignored.

        Args:
            date: The date parameter (ignored for Žabac)

        Returns:
            List of all CSV URLs available on the index page.
        """
        logger.warning(
            f"Žabac crawler ignores date parameter ({date:%Y-%m-%d}) - "
            "only current CSV files are available"
        )

        content = self.fetch_text(self.BASE_URL)

        if not content:
            logger.warning(f"No content found at Žabac index URL: {self.BASE_URL}")
            return []

        all_urls = self.parse_index(content)

        if not all_urls:
            logger.warning("No Žabac CSV URLs found on index page")

        return all_urls

    def get_all_products(self, date: datetime.date) -> list[Store]:
        """
        Main method to fetch and parse all Žabac store, product, and price info.

        Note: Date parameter is ignored as Žabac only provides current prices.

        Args:
            date: The date parameter (ignored for Žabac)

        Returns:
            List of Store objects with their products.
        """
        csv_links = self.get_index(date)

        if not csv_links:
            logger.warning("No Žabac CSV links found")
            return []

        stores = []
        for url in csv_links:
            try:
                store = self.parse_store_info(url)
                products = self.get_store_prices(url)
            except ValueError as ve:
                logger.error(
                    f"Skipping store due to parsing error from URL {url}: {ve}",
                    exc_info=False,
                )
                continue
            except Exception as e:
                logger.error(
                    f"Error processing Žabac store from {url}: {e}", exc_info=True
                )
                continue

            if not products:
                logger.warning(f"No products found for Žabac store at {url}, skipping.")
                continue

            store.items = products
            stores.append(store)

        return stores

    def fix_product_data(self, data: dict) -> dict:
        """
        Clean and fix Žabac-specific product data.

        Args:
            data: Dictionary containing the row data

        Returns:
            The cleaned data
        """
        if "product" in data and data["product"]:
            data["product"] = data["product"].strip()

        # Call parent method for common fixups
        return super().fix_product_data(data)


if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG)
    crawler = ZabacCrawler()
    stores = crawler.crawl(datetime.date.today())
    print(stores[0])
    print(stores[0].items[0])
